from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QToolButton, QLabel, QLineEdit, QTreeWidget, QTreeWidgetItem, QMessageBox,
                             QPushButton, QInputDialog, QDialog, QDialogButtonBox, QProgressBar, QFormLayout,
                             QTableWidget, QTableWidgetItem, QComboBox, QFrame, QCalendarWidget, QProgressDialog,
                             QHeaderView, QCheckBox, QFileDialog, QAbstractItemView, QMenu, QScrollArea)
from PyQt6.QtCore import Qt, QSize, QThread, pyqtSignal, QObject, QDate, QByteArray
from PyQt6.QtGui import QIcon, QColor, QPixmap, QFont, QPainter, QPalette
from PyQt6.QtSvg import QSvgRenderer
from PyQt6.QtCore import Qt, QByteArray
from io import BytesIO
import pymysql
import sys
import os
import time
import sqlite3
import hashlib
from io import BytesIO
import json
import os.path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from datetime import datetime

# Constantes para la conexión a la base de datos
DB_CONFIG = {
    "charset": "utf8mb4",
    "cursorclass": pymysql.cursors.DictCursor,
    "db": "FLUJOGRAMA",
    "host": "servicioalochoro-prueba1631.l.aivencloud.com",
    "password": "AVNS_XIL6StsPZSOwo0ZxNfr",
    "port": 15140,
    "user": "avnadmin",
}

# Constantes de colores basadas en la interfaz actual
COLORS = {
    'background': '#1A1F2B',     # Azul muy oscuro del fondo
    'surface': '#242B3D',        # Azul oscuro de los paneles
    'text': '#FFFFFF',           # Blanco para el texto
    'primary': '#0066B3',        # Azul del logo
    'primary_dark': '#004C8C',   # Azul oscuro para hover
    'primary_light': '#3399FF',  # Azul claro para elementos activos
    'accent': '#FF6B35',         # Naranja del logo
    'button_bg': '#2A324A',      # Color de fondo de los botones
    'button_hover': '#343D5C',   # Color hover de los botones
    'error': '#FF4444',          # Rojo para errores
    'success': '#4CAF50',        # Verde para éxito
    'text_secondary': '#8A93A7'  # Gris azulado para texto secundario
}


# Función para crear un icono desde SVG
def create_icon_from_svg(svg_content, color='white'):
    # Reemplazar el color en el SVG
    svg_content = svg_content.replace('currentColor', color)
    
    # Crear QPixmap
    pixmap = QPixmap(24, 24)
    pixmap.fill(Qt.GlobalColor.transparent)
    
    # Crear QIcon directamente desde el SVG como imagen
    icon = QIcon()
    icon.addPixmap(pixmap)
    
    return icon

def resource_path(relative_path):
    """Obtiene la ruta absoluta del recurso, funciona para desarrollo y PyInstaller"""
    base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)

class WorkerSignals(QObject):
    finished = pyqtSignal(bool, str)
    progress = pyqtSignal(int)

class WorkerThread(QThread):
    def __init__(self, function, *args, **kwargs):
        super().__init__()
        self.function = function
        self.args = args
        self.kwargs = kwargs
        self.signals = WorkerSignals()
    def run(self):
        try:
            result = self.function(self.signals.progress, *self.args, **self.kwargs)
            self.signals.finished.emit(True, "Operación completada con éxito")
        except Exception as e:
            self.signals.finished.emit(False, str(e))
class DatabaseManager:
    _connection_pool = None
    _max_retries = 3
    _retry_delay = 1  # segundos
    
    @classmethod
    def get_connection(cls):
        """Obtiene una conexión del pool"""
        if cls._connection_pool is None:
            cls._connection_pool = pymysql.connect(**DB_CONFIG)
        return cls._connection_pool

    @classmethod
    def execute_query(cls, query, params=None, retries=0):
        """Ejecuta una query con reintentos y manejo de errores mejorado"""
        try:
            connection = cls.get_connection()
            with connection.cursor() as cursor:
                cursor.execute(query, params)
                
                if query.strip().upper().startswith('SELECT'):
                    result = cursor.fetchall()
                    return result
                    
                connection.commit()
                return cursor.rowcount
                
        except pymysql.Error as e:
            if retries < cls._max_retries:
                time.sleep(cls._retry_delay)
                return cls.execute_query(query, params, retries + 1)
            raise Exception(f"Error en la base de datos después de {cls._max_retries} intentos: {str(e)}")
            
        except Exception as e:
            raise Exception(f"Error inesperado: {str(e)}")

    @classmethod
    def batch_insert(cls, table, columns, values):
        """Inserta múltiples registros en una sola transacción"""
        try:
            placeholders = ', '.join(['%s'] * len(columns))
            query = f"""
                INSERT INTO {table} 
                ({', '.join(columns)}) 
                VALUES ({placeholders})
            """
            
            connection = cls.get_connection()
            with connection.cursor() as cursor:
                cursor.executemany(query, values)
                connection.commit()
                return cursor.rowcount
                
        except Exception as e:
            raise Exception(f"Error en inserción por lotes: {str(e)}")

    @classmethod
    def execute_transaction(cls, queries):
        """Ejecuta múltiples queries en una transacción"""
        connection = cls.get_connection()
        try:
            with connection.cursor() as cursor:
                for query, params in queries:
                    cursor.execute(query, params)
                connection.commit()
                
        except Exception as e:
            connection.rollback()
            raise Exception(f"Error en transacción: {str(e)}")

    @staticmethod
    def get_last_id():
        result = DatabaseManager.execute_query("SELECT MAX(id) as max_id FROM documento")
        return result[0]['max_id'] if result else 0

    @staticmethod
    def init_user_table():
        query = """
        CREATE TABLE IF NOT EXISTS usuarios (
            id INT AUTO_INCREMENT PRIMARY KEY,
            username VARCHAR(50) UNIQUE NOT NULL,
            password VARCHAR(255) NOT NULL,
            email VARCHAR(100) UNIQUE NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
        DatabaseManager.execute_query(query)

    @staticmethod
    def register_user(username, email, password, rol="usuario", departamento=None, fecha_solicitud=None):
        """Registra un nuevo usuario con email y departamento"""
        try:
            salt = DatabaseManager.generate_salt()
            password_hash = DatabaseManager.hash_password(password, salt)
            
            # Sentencia SQL actualizada con departamento
            query = """
            INSERT INTO usuario 
                (nombreusuario, email, password_hash, salt, rol, departamento, fecha_solicitud) 
            VALUES 
                (%s, %s, %s, %s, %s, %s, CONVERT_TZ(NOW(), 'UTC', 'America/Santiago'))
            """
            
            # Parámetros actualizados incluyendo departamento
            params = (username, email, password_hash, salt, rol, departamento)
            
            # Ejecutar la consulta
            DatabaseManager.execute_query(query, params)
            return True
            
        except pymysql.err.IntegrityError as e:
            if "Duplicate entry" in str(e):
                if "email" in str(e):
                    raise Exception("El email ya está registrado en el sistema")
                else:
                    raise Exception("El nombre de usuario ya está registrado en el sistema")
            raise Exception(f"Error de integridad en la base de datos: {str(e)}")
        except Exception as e:
            raise Exception(f"Error al registrar usuario: {str(e)}")

    @staticmethod
    def validate_login(email, password):
        """Modificado para usar email en lugar de nombreusuario"""
        query = """
        SELECT password_hash, salt, rol, departamento 
        FROM usuario 
        WHERE email = %s
        """
        result = DatabaseManager.execute_query(query, (email,))
        
        if result:
            user = result[0]
            password_hash = DatabaseManager.hash_password(password, user['salt'])
            if password_hash == user['password_hash']:
                return True, user['rol'], user['departamento']
        return False, None, None

    @staticmethod
    def generate_salt():
        return os.urandom(32).hex()

    @staticmethod
    def hash_password(password, salt):
        return hashlib.sha256((password + salt).encode()).hexdigest()

    @staticmethod
    def get_establecimientos():
        try:
            query = "SELECT nombre_establecimiento FROM establecimiento ORDER BY nombre_establecimiento"
            resultados = DatabaseManager.execute_query(query)
            return [resultado['nombre_establecimiento'] for resultado in resultados]
        except Exception as e:
            print(f"Error al obtener establecimientos: {str(e)}")
            return []

    @staticmethod
    def get_departamentos():
        try:
            query = "SELECT nombre_departamento FROM departamento ORDER BY id_departamento"
            resultados = DatabaseManager.execute_query(query)
            return [resultado['nombre_departamento'] for resultado in resultados]
        except Exception as e:
            print(f"Error al obtener departamentos: {str(e)}")
            return []

class ProgressDialog(QDialog):
    def __init__(self, parent=None, title="Progreso", description="Procesando..."):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setFixedSize(300, 100)
        layout = QVBoxLayout(self)
        
        self.description_label = QLabel(description)
        layout.addWidget(self.description_label)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid grey;
                border-radius: 5px;
                text-align: center;
            }

            QProgressBar::chunk {
                background-color: #4CAF50;
                width: 10px;
                margin: 0.5px;
            }
        """)
        layout.addWidget(self.progress_bar)
        
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowStaysOnTopHint)
        self.setModal(True)

    def set_range(self, minimum, maximum):
        self.progress_bar.setRange(minimum, maximum)

    def set_value(self, value):
        self.progress_bar.setValue(value)

class MainWindow(QMainWindow):
    def __init__(self, email=None, user_role=None, departamento=None):
        super().__init__()
        self.email = email
        self.user_role = user_role
        self.departamento = departamento
        self.init_ui()
        self.setup_button_visibility()
        self.setup_user_info()
        
    def setup_user_info(self):
        """Configura la información del usuario en la interfaz"""
        try:
            # Buscar el QLabel que muestra la información del usuario
            user_info = self.findChild(QLabel, "user_info")
            if user_info:
                # Actualizar el texto con la información del usuario actual
                user_info.setText(f"Usuario: {self.email}\nRol: {self.user_role}\nDepartamento: {self.departamento}")  # Agregado departamento
                user_info.setStyleSheet(f"""
                    QLabel {{
                        color: white;
                        background-color: {COLORS['surface']};
                        padding: 10px;
                        border-radius: 8px;
                        font-size: 13px;
                    }}
                """)
        except Exception as e:
            print(f"Error al configurar información del usuario: {e}")

    def setup_button_visibility(self):
        """Configura la visibilidad de los botones según el rol del usuario"""
        # Mapeo de roles y sus permisos
        role_permissions = {
            "admin": ["Agregar Nuevo Documento", "Consultar Documento", 
                     "Eliminar Documento", "Modificar Documento", 
                     "Administrar", "📥 Recibir Documentos", "📊 Generar Reporte Actual"],  # Agregado a permisos de admin
            "recepcionista": ["Agregar Nuevo Documento", "Consultar Documento",
                            "📥 Recibir Documentos", "📊 Generar Reporte Actual"],  # Agregado a permisos de recepcionista
            "usuario": ["Consultar Documento", "📥 Recibir Documentos", "📊 Generar Reporte Actual"]  # Agregado a permisos de usuario
        }
        
        # Obtener los permisos para el rol actual
        allowed_buttons = role_permissions.get(self.user_role.lower(), [])
        
        # Añadir "Cerrar Sesión" a los botones permitidos para todos los roles
        allowed_buttons.append("Cerrar Sesión")
        
        # Configurar visibilidad de botones
        for button in self.findChildren(QPushButton):
            if button.text() in allowed_buttons:
                button.setVisible(True)
            else:
                button.setVisible(False)

    def init_ui(self):
        self.setWindowTitle("Interfaz de Corporación Isla de Maipo")
        self.setGeometry(0, 0, 1920, 1080)
        self.setWindowIcon(QIcon(resource_path("isla_de_maipo.png")))
        
        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Panel izquierdo
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setSpacing(15)
        left_layout.setContentsMargins(10, 20, 10, 20)

        # Logo (fijo)
        logo_label = QLabel()
        logo_pixmap = QPixmap(resource_path("isla_de_maipo.png"))
        scaled_pixmap = logo_pixmap.scaled(250, 180, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
        logo_label.setPixmap(scaled_pixmap)
        logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        left_layout.addWidget(logo_label)

        # Área de scroll solo para los botones
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        
        # Contenedor de botones
        buttons_container = QWidget()
        buttons_layout = QVBoxLayout(buttons_container)
        buttons_layout.setSpacing(10)
        buttons_layout.setContentsMargins(0, 0, 0, 0)

        # Botones principales con scroll
        self.buttons_data = [
            ("Agregar Nuevo Documento", self.agregar_datos),
            ("Consultar Documento", self.consultar_datos),
            ("Eliminar Documento", self.eliminar_datos),
            ("Modificar Documento", self.modificar_datos),
            ("Administrar", self.show_admin_panel),
            ("📬 Recibir Documentos", self.recibir_documento),
            ("📊 Generar Reporte Actual", self.generar_reporte_actual)
        ]

        for text, slot in self.buttons_data:
            btn = QPushButton(text)
            btn.setObjectName(text)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {COLORS['surface']};
                    color: white;
                    border: none;
                    padding: 15px;
                    border-radius: 8px;
                    font-size: 14px;
                    font-weight: bold;
                    text-align: left;
                    padding-left: 20px;
                }}
                QPushButton:hover {{
                    background-color: {COLORS['primary']};
                }}
                QPushButton:pressed {{
                    background-color: {COLORS['primary_dark']};
                }}
            """)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(slot)
            buttons_layout.addWidget(btn)

        buttons_layout.addStretch()
        
        # Configurar el área de scroll
        scroll_area.setWidget(buttons_container)
        scroll_area.setStyleSheet(f"""
            QScrollArea {{
                border: none;
                background-color: {COLORS['background']};
            }}
            QScrollBar:vertical {{
                background-color: {COLORS['surface']};
                width: 10px;
                margin: 0px;
            }}
            QScrollBar::handle:vertical {{
                background-color: {COLORS['primary']};
                min-height: 30px;
                border-radius: 5px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
        """)
        
        # Agregar el área de scroll al layout principal
        left_layout.addWidget(scroll_area)

        # Información de usuario (fijo)
        user_info = QLabel(f"Usuario: {self.email}\nRol: {self.user_role}\nDepartamento: {self.departamento}")
        user_info.setObjectName("user_info")
        user_info.setStyleSheet(f"""
            QLabel {{
                color: white;
                background-color: {COLORS['surface']};
                padding: 10px;
                border-radius: 8px;
                font-size: 13px;
            }}
        """)
        left_layout.addWidget(user_info)

        # Botón de cerrar sesión (fijo)
        logout_btn = QPushButton("Cerrar Sesión")
        logout_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['error']};
                color: white;
                border: none;
                padding: 10px;
                border-radius: 6px;
                font-size: 13px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #d32f2f;
            }}
        """)
        logout_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        logout_btn.clicked.connect(self.logout)
        left_layout.addWidget(logout_btn)

        main_layout.addWidget(left_panel, 1)

        # Panel derecho (resto del código del panel derecho permanece igual)
        right_panel = QWidget()
        right_panel.setObjectName("right_panel")
        right_panel.setStyleSheet(f"""
            #right_panel {{
                background-color: {COLORS['background']};
                padding: 20px;  /* Padding interior */
                border-radius: 12px;
            }}
        """)

        # Layout para el panel derecho
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(20, 20, 20, 20)  # Márgenes externos
        right_layout.setSpacing(10)  # Espacio entre widgets

        # Contenedor para la barra de búsqueda y botones
        search_container = QWidget()
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(0, 0, 0, 0)
        search_layout.setSpacing(10)

        # Modificar el combo de búsqueda para categorías de agrupación
        self.search_combo = QComboBox()
        self.search_combo.addItems([
            "Id",
            "Establecimiento",
            "Tipo Doc",
            "Estado",
            "Destino"
        ])
        self.search_combo.setStyleSheet(f"""
            QComboBox {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                padding: 8px;
                border: 2px solid {COLORS['primary']};
                border-radius: 6px;
                min-width: 150px;
            }}
        """)
        self.search_combo.currentTextChanged.connect(self.agrupar_datos)

        # Barra de búsqueda (ahora independiente)
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("Buscar en todos los campos...")
        self.search_bar.setStyleSheet(f"""
            QLineEdit {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                padding: 8px 12px;
                border: 2px solid {COLORS['primary']};
                border-radius: 6px;
                font-size: 14px;
            }}
        """)
        self.search_bar.textChanged.connect(self.filtrar_busqueda)

        # Botón limpiar búsqueda
        clear_btn = QPushButton("✕")
        clear_btn.setFixedSize(32, 32)
        clear_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                border: 2px solid {COLORS['primary']};
                border-radius: 16px;
                font-size: 14px;
            }}
        """)
        clear_btn.clicked.connect(self.clear_search)

        # Agregar widgets al layout
        search_layout.addWidget(self.search_combo)
        search_layout.addWidget(self.search_bar)
        search_layout.addWidget(clear_btn)

        right_layout.addWidget(search_container)

        # TreeWidget para mostrar datos
        self.tree_widget = QTreeWidget()
        self.tree_widget.setHeaderLabels([
            "ID", "Fecha", "Establecimiento", "Tipo Doc", 
            "Nro Doc", "Materia", "Destino", "Firma", "Estado"
        ])
        self.tree_widget.setStyleSheet(f"""
            QTreeWidget {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                border: none;
                border-radius: 8px;
                gridline-color: rgba(255, 255, 255, 0.1);
                outline: none;
            }}
            
            QTreeWidget::item {{
                height: 30px;  /* Altura reducida de las filas */
                background-color: {COLORS['background']};
                border-bottom: 1px solid rgba(255, 255, 255, 0.1);
                margin: 1px;  /* Margen reducido entre filas */
                border-radius: 2px;
                padding: 0px 5px;  /* Padding horizontal reducido */
                font-size: 12px;  /* Tamaño de fuente más pequeño */
            }}
            
            QTreeWidget::item:selected {{
                background-color: {COLORS['button_bg']};
                color: {COLORS['text']};
            }}
            
            QTreeWidget::item:hover {{
                background-color: {COLORS['button_hover']};
            }}
            
            QHeaderView::section {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                padding: 5px;  /* Padding reducido en el encabezado */
                border: none;
                border-bottom: 1px solid rgba(255, 255, 255, 0.2);
                font-weight: bold;
                font-size: 12px;  /* Tamaño de fuente del encabezado */
            }}
            
            QScrollBar:vertical {{
                background-color: {COLORS['surface']};
                width: 14px;
                margin: 0px;
            }}
            
            QScrollBar::handle:vertical {{
                background-color: {COLORS['button_bg']};
                min-height: 30px;
                border-radius: 7px;
                margin: 2px;
            }}
            
            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
            
            QScrollBar:horizontal {{
                background-color: {COLORS['surface']};
                height: 14px;
                margin: 0px;
            }}
            
            QScrollBar::handle:horizontal {{
                background-color: {COLORS['button_bg']};
                min-width: 30px;
                border-radius: 7px;
                margin: 2px;
            }}
            
            QScrollBar::add-line:horizontal,
            QScrollBar::sub-line:horizontal {{
                width: 0px;
            }}
        """)
        # Configuración adicional del TreeWidget
        self.tree_widget.setAlternatingRowColors(False)  # Desactivar colores alternados
        self.tree_widget.setIndentation(0)  # Eliminar la indentación
        self.tree_widget.setRootIsDecorated(False)  # Ocultar los triángulos expandibles
        self.tree_widget.setUniformRowHeights(True)  # Altura uniforme para las filas
        self.tree_widget.setVerticalScrollMode(QTreeWidget.ScrollMode.ScrollPerPixel)
        self.tree_widget.setHorizontalScrollMode(QTreeWidget.ScrollMode.ScrollPerPixel)

        # Configurar el header
        header = self.tree_widget.header()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        header.setStretchLastSection(True)
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        # Configurar las columnas para que coincidan con la imagen
        self.tree_widget.setHeaderLabels([
            "ID", "Fecha", "Establecimiento", "Tipo Doc", 
            "Nro Doc", "Materia", "Destino", "Firma", "Estado", "PDF"
        ])

        # Ajustar el ancho de las columnas
        column_widths = {
            0: 40,   # ID
            1: 80,   # Fecha
            2: 120,  # Establecimiento
            3: 80,   # Tipo Doc
            4: 70,   # Nro Doc
            5: 120,  # Materia
            6: 100,  # Destino
            7: 60,   # Firma
            8: 70,   # Estado
            9: 40    # PDF
        }

        # Aplicar los anchos de columna
        for col, width in column_widths.items():
            self.tree_widget.setColumnWidth(col, width)
        
        right_layout.addWidget(self.tree_widget)

        # Agregar el panel derecho al layout principal
        main_layout.addWidget(right_panel, 4)

        # Estilo general de la ventana
        self.setStyleSheet(f"""
            QMainWindow {{
                background-color: {COLORS['background']};
            }}
            QWidget {{
                background-color: {COLORS['background']};
                color: {COLORS['text']};
            }}
            QPushButton {{
                background-color: {COLORS['button_bg']};
                color: {COLORS['text']};
                border: none;
                padding: 15px;
                border-radius: 8px;
                font-size: 14px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['button_hover']};
            }}
            QLineEdit {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                border: none;
                border-radius: 4px;
                padding: 8px;
            }}
            QTreeWidget {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                border: none;
                border-radius: 8px;
            }}
            QTreeWidget::item {{
                padding: 8px;
            }}
            QTreeWidget::item:selected {{
                background-color: {COLORS['button_bg']};
            }}
            QHeaderView::section {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                padding: 8px;
                border: none;
            }}
            QComboBox {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                border: none;
                border-radius: 4px;
                padding: 8px;
            }}
            QComboBox:hover {{
                background-color: {COLORS['button_hover']};
            }}
            QLabel {{
                color: {COLORS['text']};
            }}
            #left_panel {{
                background-color: {COLORS['surface']};
                border-radius: 12px;
            }}
        """)

    def agregar_datos(self):
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle("Agregar Documento")
            dialog.setFixedWidth(400)
            layout = QVBoxLayout(dialog)

            # Crear formulario
            form_layout = QFormLayout()
            inputs = {}

            # Fecha
            fecha_input = QCalendarWidget()
            fecha_input.setSelectedDate(QDate.currentDate())
            form_layout.addRow("Fecha:", fecha_input)

            # Campos de texto y sus valores
            campos = [
                ("establecimiento", QComboBox(), DatabaseManager.get_establecimientos()),
                ("tipodocumento", QComboBox(), [
                    "Oficio", "Resolucion", "Ordinario", "Memo", 
                    "Decreto", "Factura", "Carta"
                ]),
                ("nrodocumento", QLineEdit(), None),
                ("materia", QLineEdit(), None),
                ("firma", QLineEdit(), None),
                ("estado", QLineEdit(), None)
            ]

            for campo, widget, opciones in campos:
                if isinstance(widget, QComboBox):
                    if opciones:
                        widget.addItems(opciones)
                inputs[campo] = widget
                form_layout.addRow(f"{campo.capitalize()}:", widget)

            # Botón para agregar PDF
            pdf_button = QPushButton("Agregar PDF")
            pdf_button.setStyleSheet(f"""
                QPushButton {{
                    background-color: {COLORS['primary']};
                    color: {COLORS['text']};
                    padding: 8px;
                    border: none;
                    border-radius: 4px;
                }}
                QPushButton:hover {{
                    background-color: {COLORS['primary_light']};
                }}
            """)
            
            pdf_content = [None]  # Lista para almacenar el contenido del PDF
            
            def seleccionar_pdf():
                try:
                    file_name, _ = QFileDialog.getOpenFileName(
                        dialog,
                        "Seleccionar PDF",
                        "",
                        "PDF Files (*.pdf)"
                    )
                    if file_name:
                        with open(file_name, 'rb') as file:
                            pdf_content[0] = file.read()
                        pdf_button.setText("PDF Seleccionado ✓")
                        pdf_button.setStyleSheet(f"""
                            QPushButton {{
                                background-color: {COLORS['success']};
                                color: {COLORS['text']};
                                padding: 8px;
                                border: none;
                                border-radius: 4px;
                            }}
                        """)
                except Exception as e:
                    self.mostrar_mensaje(
                        "Error",
                        f"Error al seleccionar PDF: {str(e)}",
                        QMessageBox.Icon.Critical
                    )

            pdf_button.clicked.connect(seleccionar_pdf)
            form_layout.addRow("PDF:", pdf_button)

            layout.addLayout(form_layout)

            # Botones
            button_box = QDialogButtonBox(
                QDialogButtonBox.StandardButton.Save | 
                QDialogButtonBox.StandardButton.Cancel
            )

            def guardar_documento():
                try:
                    # Validar campos
                    for campo, widget in inputs.items():
                        if isinstance(widget, QLineEdit) and not widget.text().strip():
                            raise ValueError(f"El campo {campo} no puede estar vacío")
                        elif isinstance(widget, QComboBox) and not widget.currentText():
                            raise ValueError(f"Debe seleccionar un {campo}")

                    # Construir la consulta
                    query = """
                        INSERT INTO documento 
                        (fecha, establecimiento, tipodocumento, nrodocumento, 
                         materia, firma, estado, lugar_actual, archivo_pdf)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    
                    valores = [
                        fecha_input.selectedDate().toString("yyyy-MM-dd"),
                        inputs['establecimiento'].currentText(),
                        inputs['tipodocumento'].currentText(),
                        inputs['nrodocumento'].text(),
                        inputs['materia'].text(),
                        inputs['firma'].text(),
                        inputs['estado'].text(),
                        self.departamento,  # Usar el departamento actual del usuario
                        pdf_content[0]  # Contenido del PDF
                    ]

                    DatabaseManager.execute_query(query, valores)

                    self.mostrar_mensaje(
                        "Éxito",
                        "Documento agregado correctamente",
                        QMessageBox.Icon.Information
                    )

                    dialog.accept()
                    self.consultar_datos()

                except ValueError as e:
                    self.mostrar_mensaje(
                        "Error de validación",
                        str(e),
                        QMessageBox.Icon.Warning
                    )
                except Exception as e:
                    self.mostrar_mensaje(
                        "Error",
                        f"Error al guardar documento: {str(e)}",
                        QMessageBox.Icon.Critical
                    )

            button_box.accepted.connect(guardar_documento)
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)

            # Aplicar estilos
            dialog.setStyleSheet(f"""
                QDialog {{
                    background-color: {COLORS['background']};
                }}
                QLabel {{
                    color: {COLORS['text']};
                }}
                QLineEdit, QComboBox {{
                    background-color: {COLORS['surface']};
                    color: {COLORS['text']};
                    padding: 8px;
                    border: 1px solid {COLORS['primary']};
                    border-radius: 4px;
                }}
            """)

            dialog.exec()

        except Exception as e:
            self.mostrar_mensaje(
                "Error",
                f"Error al abrir formulario: {str(e)}",
                QMessageBox.Icon.Critical
            )

    def guardar_datos_seguro(self):
        try:
            data = {key: entry.text() for key, entry in self.entries.items()}
            if not all(data.values()):
                raise ValueError("Todos los campos deben estar llenos")
            
            DatabaseManager.execute_query(                """INSERT INTO documento(
                    fecha, establecimiento, tipodocumento, 
                    nrodocumento, materia, destino, firma, estado
                ) VALUES(%s, %s, %s, %s, %s, %s,                         %s, %s)""",
                (data['fecha'],
                 data['establecimiento'],
                 data['tipodocumento'],
                 data['nrodocumento'],
                 data['materia'],
                 data['destino'],
                 data['firma'],
                 data['estado']
                ))
            for entry in self.entries.values():
                entry.clear()
            
            self.mostrar_mensaje("Éxito", "Datos agregados exitosamente")
        except ValueError as e:
            self.mostrar_mensaje("Error de validación", str(e), QMessageBox.Icon.Warning)
        except Exception as e:
            self.mostrar_mensaje("Error inesperado", f"Ocurrió un error al guardar los datos: {str(e)}", QMessageBox.Icon.Critical)

    def consultar_datos(self):
        try:
            # Limpiar datos anteriores
            self.tree_widget.clear()
            
            # Cerrar conexión existente si hay una
            if hasattr(DatabaseManager, '_connection_pool') and DatabaseManager._connection_pool:
                DatabaseManager._connection_pool.close()
                DatabaseManager._connection_pool = None
                
            # Obtener nueva conexión
            connection = DatabaseManager.get_connection()
            
            # Consulta SQL actualizada para incluir lugar_actual
            query = """
                SELECT 
                    id_documento,
                    fecha,
                    establecimiento,
                    tipodocumento,
                    nrodocumento,
                    materia,
                    lugar_actual,
                    destino,
                    firma,
                    estado,
                    CASE WHEN archivo_pdf IS NOT NULL THEN 1 ELSE 0 END as tiene_pdf
                FROM documento 
                ORDER BY id_documento DESC
            """
            resultados = DatabaseManager.execute_query(query)
            
            # Configurar las columnas si no están configuradas
            if self.tree_widget.columnCount() != 12:  # Ahora son 12 columnas (11 + Enviar)
                self.tree_widget.setHeaderLabels([
                    "ID", "Fecha", "Establecimiento", "Tipo Doc", 
                    "Nro Doc", "Materia", "Lugar Actual", "Destino", 
                    "Firma", "Estado", "PDF", "Enviar"
                ])
            
            # Procesar cada resultado
            for registro in resultados:
                item = QTreeWidgetItem()
                
                # Establecer los textos de las columnas
                for i, campo in enumerate(['id_documento', 'fecha', 'establecimiento', 
                                        'tipodocumento', 'nrodocumento', 'materia', 
                                        'lugar_actual', 'destino', 'firma', 'estado']):
                    item.setText(i, str(registro[campo]))
                
                # Añadir el item al tree widget
                self.tree_widget.addTopLevelItem(item)
                
                # Configurar el botón PDF en la columna PDF
                self.setup_pdf_button(item, registro['tiene_pdf'], registro['id_documento'])

                # Agregar botón de envío solo si el documento está en el departamento del usuario
                if registro['lugar_actual'] == self.departamento:
                    enviar_btn = QPushButton("📨")
                    enviar_btn.setToolTip("Enviar Documento")
                    enviar_btn.setStyleSheet(f"""
                        QPushButton {{
                            background-color: {COLORS['primary']};
                            border: none;
                            border-radius: 4px;
                            padding: 5px;
                            color: white;
                            font-size: 16px;
                        }}
                        QPushButton:hover {{
                            background-color: {COLORS['primary_dark']};
                        }}
                    """)
                    enviar_btn.clicked.connect(lambda checked, doc=registro: self.enviar_documento(doc))
                    self.tree_widget.setItemWidget(item, 11, enviar_btn)

            # Ajustar el ancho de las columnas
            column_widths = {
                0: 40,   # ID
                1: 80,   # Fecha
                2: 120,  # Establecimiento
                3: 80,   # Tipo Doc
                4: 70,   # Nro Doc
                5: 120,  # Materia
                6: 100,  # Lugar Actual
                7: 100,  # Destino
                8: 60,   # Firma
                9: 70,   # Estado
                10: 40,  # PDF
                11: 50   # Enviar
            }
            
            # Aplicar los anchos de columna
            for col, width in column_widths.items():
                self.tree_widget.setColumnWidth(col, width)
            
            # Mostrar mensaje de éxito
            QMessageBox.information(
                self,
                "Consulta exitosa",
                f"Se cargaron {len(resultados)} registros"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error al consultar datos: {str(e)}"
            )

    def actualizar_datos_sin_progreso(self):
        try:
            resultados = DatabaseManager.execute_query("SELECT * FROM documento")
            self.tree_widget.clear()
            for registro in resultados:
                item = QTreeWidgetItem(self.tree_widget)
                item.setText(0, str(registro['id']))
                item.setText(1, registro['nombre'])
                item.setText(2, str(registro['niveldemamador']))
        except Exception as e:
            self.mostrar_mensaje("Error", f"No se pudieron actualizar los datos: {str(e)}", QMessageBox.Icon.Critical)

    def closeEvent(self, event):
        """Maneja el cierre de la aplicación"""
        try:
            # Verificar si hay un worker activo
            if hasattr(self, 'worker') and self.worker is not None:
                # Detener el worker si está ejecutándose
                if self.worker.isRunning():
                    self.worker.terminate()
                    self.worker.wait()
            
            # Cerrar la conexión a la base de datos
            if hasattr(DatabaseManager, '_connection_pool') and DatabaseManager._connection_pool:
                DatabaseManager._connection_pool.close()
            
            # Guardar configuraciones
            self.save_window_state()
            
            # Aceptar el evento de cierre
            event.accept()
            
        except Exception as e:
            print(f"Error al cerrar la aplicación: {e}")
            event.accept()

    def setup_pdf_button(self, item, tiene_pdf, doc_id):
        """Configurar el botón de PDF para un item"""
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        download_btn = QPushButton("📥")
        download_btn.setFixedSize(18, 18)
        download_btn.setToolTip("Descargar PDF")
        download_btn.setEnabled(tiene_pdf)
        
        COLOR_PDF = "#FF8C00"
        COLOR_NO_PDF = "#808080"
        COLOR_HOVER = "#FF6B00"
        
        download_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLOR_PDF if tiene_pdf else COLOR_NO_PDF};
                color: white;
                border: none;
                border-radius: 9px;
                padding: 0px;
                font-size: 10px;
                qproperty-alignment: AlignCenter;
            }}
            QPushButton:hover {{
                background-color: {COLOR_HOVER if tiene_pdf else COLOR_NO_PDF};
            }}
            QPushButton:disabled {{
                background-color: {COLOR_NO_PDF};
                color: #CCCCCC;
            }}
        """)
        
        if tiene_pdf:
            download_btn.clicked.connect(
                lambda checked, x=doc_id: self.descargar_pdf(x)
            )
        
        layout.addWidget(download_btn, 0, Qt.AlignmentFlag.AlignCenter)
        self.tree_widget.setItemWidget(item, 10, container)  # Columna PDF

    def agregar_registro_al_tree(self, registro, parent=None):
        """Método auxiliar para agregar un registro al TreeWidget"""
        item = QTreeWidgetItem(parent if parent else self.tree_widget)
        
        # Establecer valores en las columnas
        campos = ['id_documento', 'fecha', 'establecimiento', 'tipodocumento', 
                 'nrodocumento', 'materia', 'destino', 'firma', 'estado']
        
        for i, campo in enumerate(campos):
            valor = registro[campo]
            if campo == 'fecha':
                valor = str(valor)
            item.setText(i, str(valor))
        
        # Crear y configurar el botón de PDF
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        download_btn = QPushButton("📥")
        download_btn.setFixedSize(18, 18)
        download_btn.setToolTip("Descargar PDF")
        
        tiene_pdf = registro.get('tiene_pdf', False)
        download_btn.setEnabled(tiene_pdf)
        
        # Colores y estilos
        COLOR_PDF = "#FF8C00"
        COLOR_NO_PDF = "#808080"
        COLOR_HOVER = "#FF6B00"
        
        download_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLOR_PDF if tiene_pdf else COLOR_NO_PDF};
                color: white;
                border: none;
                border-radius: 9px;
                padding: 0px;
                font-size: 10px;
                qproperty-alignment: AlignCenter;
            }}
            QPushButton:hover {{
                background-color: {COLOR_HOVER if tiene_pdf else COLOR_NO_PDF};
            }}
            QPushButton:disabled {{
                background-color: {COLOR_NO_PDF};
                color: #CCCCCC;
            }}
        """)
        
        if tiene_pdf:
            doc_id = registro['id_documento']
            download_btn.clicked.connect(
                lambda checked, x=doc_id: self.descargar_pdf(x)
            )
        
        layout.addWidget(download_btn, 0, Qt.AlignmentFlag.AlignCenter)
        
        # Establecer el widget en la última columna
        if parent:
            self.tree_widget.setItemWidget(item, 9, container)
        else:
            self.tree_widget.setItemWidget(item, 9, container)

    def show_admin_panel(self):
        dialog = AdminPanel(self)
        dialog.exec()

    def logout(self):
        """Función para manejar el cierre de sesión"""
        self.hide()
        
        login = LoginDialog()
        if login.exec() == QDialog.DialogCode.Accepted:
            # Actualizar las credenciales del usuario
            self.email = login.email_input.text()  # Cambiado de username a email
            self.user_role = login.get_user_role()
            self.departamento = login.get_user_departamento()  # Obtener el departamento
            
            # Actualizar la información del usuario
            self.setup_user_info()
            
            # Actualizar la visibilidad de los botones según el nuevo rol
            self.setup_button_visibility()
            
            self.show()
        else:
            QApplication.instance().quit()

    def setup_button_visibility(self):
        """Configura la visibilidad de los botones según el rol del usuario"""
        role_permissions = {
            "admin": ["Agregar Nuevo Documento", "Consultar Documento", 
                      "Eliminar Documento", "Modificar Documento", 
                      "Administrar", "📬 Recibir Documentos", "📊 Generar Reporte Actual"],
            "recepcionista": ["Agregar Nuevo Documento", "Consultar Documento",
                              "📬 Recibir Documentos", "📊 Generar Reporte Actual"],
            "usuario": ["Consultar Documento", "📬 Recibir Documentos", "📊 Generar Reporte Actual"]
        }
        
        allowed_buttons = role_permissions.get(self.user_role.lower(), [])
        allowed_buttons.append("Cerrar Sesión")
        
        for button in self.findChildren(QPushButton):
            if button.text() in allowed_buttons:
                button.setVisible(True)
            else:
                button.setVisible(False)

    def agrupar_datos(self):
        """Agrupa los datos según la categoría seleccionada"""
        categoria = self.search_combo.currentText()
        
        # Obtener todos los items y sus datos asociados
        items_data = []
        for i in range(self.tree_widget.topLevelItemCount()):
            item = self.tree_widget.takeTopLevelItem(0)
            id_doc = int(item.text(0))
            
            # Obtener información completa del documento
            query = """
                SELECT 
                    archivo_pdf IS NOT NULL AND archivo_pdf != '' as tiene_pdf,
                    tipodocumento,
                    lugar_actual,
                    destino,
                    nrodocumento
                FROM documento 
                WHERE id_documento = %s
            """
            resultado = DatabaseManager.execute_query(query, (id_doc,))
            doc_info = resultado[0] if resultado else None
            
            items_data.append({
                'item': item,
                'tiene_pdf': doc_info['tiene_pdf'] if doc_info else False,
                'id_documento': id_doc,
                'lugar_actual': doc_info['lugar_actual'] if doc_info else '',
                'tipodocumento': doc_info['tipodocumento'] if doc_info else '',
                'destino': doc_info['destino'] if doc_info else ''
            })
        
        # Mapeo de categorías a índices de columna
        columnas = {
            "Id": 0,
            "Establecimiento": 2,
            "Tipo Doc": 3,
            "Estado": 8,
            "Destino": 6
        }
        
        if categoria not in columnas:
            return
        
        # Agrupar items por la categoría seleccionada
        grupos = {}
        for data in items_data:
            item = data['item']
            key = item.text(columnas[categoria])
            if key not in grupos:
                grupos[key] = []
            grupos[key].append(data)
        
        # Limpiar y repoblar el tree widget
        self.tree_widget.clear()
        
        # Ordenar las claves de los grupos en orden descendente
        if categoria == "Id":
            sorted_keys = sorted(grupos.keys(), key=lambda x: int(x), reverse=True)
        else:
            sorted_keys = sorted(grupos.keys(), reverse=True)
        
        # Agregar items agrupados en orden
        for grupo in sorted_keys:
            grupo_items = grupos[grupo]
            # Ordenar items dentro del grupo por ID en orden descendente
            grupo_items.sort(key=lambda x: int(x['item'].text(0)), reverse=True)
            for data in grupo_items:
                item = data['item']
                self.tree_widget.addTopLevelItem(item)
                self.setup_pdf_button(item, data['tiene_pdf'], data['id_documento'])
                if data['lugar_actual'] == self.departamento:
                    self.setup_enviar_button(item, data)

    def setup_enviar_button(self, item, documento):
        """Configura el botón de envío para un item"""
        enviar_btn = QPushButton("📨")
        enviar_btn.setToolTip("Enviar Documento")
        enviar_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['primary']};
                border: none;
                border-radius: 4px;
                padding: 5px;
                color: white;
                font-size: 16px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['primary_dark']};
            }}
        """)
        
        # Obtener información completa del documento
        try:
            query = """
                SELECT id_documento, tipodocumento, lugar_actual, destino, nrodocumento
                FROM documento 
                WHERE id_documento = %s
            """
            result = DatabaseManager.execute_query(query, (documento['id_documento'],))
            if result:
                doc_info = {
                    'id_documento': result[0]['id_documento'],
                    'tipodocumento': result[0]['tipodocumento'],
                    'lugar_actual': result[0]['lugar_actual'],
                    'destino': result[0]['destino'],
                    'nrodocumento': result[0]['nrodocumento']
                }
                enviar_btn.clicked.connect(lambda: self.enviar_documento(doc_info))
        except Exception as e:
            print(f"Error al obtener información del documento: {str(e)}")
            enviar_btn.setEnabled(False)
        
        self.tree_widget.setItemWidget(item, 11, enviar_btn)

    def filtrar_busqueda(self, texto):
        """Filtra los datos según el texto de búsqueda en todas las columnas"""
        texto = texto.lower()
        for i in range(self.tree_widget.topLevelItemCount()):
            item = self.tree_widget.topLevelItem(i)
            mostrar = False
            
            # Buscar en todas las columnas
            for j in range(self.tree_widget.columnCount()):
                if texto in item.text(j).lower():
                    mostrar = True
                    break
                    
            item.setHidden(not mostrar)

    def clear_search(self):
        """Limpia la búsqueda y muestra todos los registros"""
        self.search_bar.clear()
        for i in range(self.tree_widget.topLevelItemCount()):
            self.tree_widget.topLevelItem(i).setHidden(False)

    def eliminar_datos(self):
        try:
            # Verificar si hay un item seleccionado
            selected_items = self.tree_widget.selectedItems()
            if not selected_items:
                self.mostrar_mensaje(
                    "Error",
                    "Por favor seleccione un documento para eliminar",
                    QMessageBox.Icon.Warning
                )
                return

            # Obtener el ID del documento seleccionado
            id_documento = selected_items[0].text(0)

            # Confirmar eliminación
            confirmacion = QMessageBox.question(
                self,
                "Confirmar Eliminación",
                "¿Está seguro de que desea eliminar este documento?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )

            if confirmacion == QMessageBox.StandardButton.Yes:
                # Eliminar el documento
                query = "DELETE FROM documento WHERE id_documento = %s"
                DatabaseManager.execute_query(query, (id_documento,))

                self.mostrar_mensaje(
                    "Éxito",
                    "Documento eliminado correctamente",
                    QMessageBox.Icon.Information
                )

                # Actualizar la vista
                self.consultar_datos()

        except Exception as e:
            self.mostrar_mensaje(
                "Error",
                f"Error al eliminar documento: {str(e)}",
                QMessageBox.Icon.Critical
            )
            if hasattr(DatabaseManager, '_connection_pool') and DatabaseManager._connection_pool:
                DatabaseManager._connection_pool.close()
                DatabaseManager._connection_pool = None
            

    def modificar_datos(self):
        try:
            # Verificar si hay un item seleccionado
            selected_items = self.tree_widget.selectedItems()
            if not selected_items:
                self.mostrar_mensaje(
                    "Error",
                    "Por favor seleccione un documento para modificar",
                    QMessageBox.Icon.Warning
                )
                return

            # Obtener el ID del documento seleccionado
            id_documento = selected_items[0].text(0)

            # Crear diálogo de modificación
            dialog = QDialog(self)
            dialog.setWindowTitle("Modificar Documento")
            dialog.setFixedWidth(400)
            layout = QVBoxLayout(dialog)

            # Obtener datos actuales del documento
            query = """
                SELECT fecha, establecimiento, tipodocumento, nrodocumento, 
                       materia, destino, firma, estado, lugar_actual
                FROM documento 
                WHERE id_documento = %s
            """
            resultado = DatabaseManager.execute_query(query, (id_documento,))[0]

            # Crear formulario
            form_layout = QFormLayout()
            inputs = {}

            # Fecha
            fecha_input = QCalendarWidget()
            fecha_input.setSelectedDate(QDate.fromString(str(resultado['fecha']), "yyyy-MM-dd"))
            form_layout.addRow("Fecha:", fecha_input)

            # Campos de texto y sus valores actuales
            campos = [
                ("establecimiento", QComboBox(), DatabaseManager.get_establecimientos()),
                ("tipodocumento", QComboBox(), [
                    "Oficio", "Resolucion", "Ordinario", "Memo", 
                    "Decreto", "Factura", "Carta"
                ]),
                ("nrodocumento", QLineEdit(), None),
                ("materia", QLineEdit(), None),
                ("firma", QLineEdit(), None),
                ("estado", QLineEdit(), None)
            ]

            for campo, widget, opciones in campos:
                if isinstance(widget, QComboBox):
                    if opciones:
                        widget.addItems(opciones)
                    widget.setCurrentText(str(resultado[campo]))
                else:
                    widget.setText(str(resultado[campo]))
                inputs[campo] = widget
                form_layout.addRow(f"{campo.capitalize()}:", widget)

            # Agregar botón para cambiar PDF
            pdf_button = QPushButton("Cambiar PDF")
            pdf_button.setStyleSheet(f"""
                QPushButton {{
                    background-color: {COLORS['primary']};
                    color: {COLORS['text']};
                    padding: 8px;
                    border: none;
                    border-radius: 4px;
                }}
                QPushButton:hover {{
                    background-color: {COLORS['primary_light']};
                }}
            """)
            
            def cambiar_pdf():
                try:
                    file_name, _ = QFileDialog.getOpenFileName(
                        dialog,
                        "Seleccionar PDF",
                        "",
                        "PDF Files (*.pdf)"
                    )
                    if file_name:
                        with open(file_name, 'rb') as file:
                            pdf_content = file.read()
                            
                        # Actualizar el PDF en la base de datos
                        update_pdf_query = """
                            UPDATE documento 
                            SET archivo_pdf = %s
                            WHERE id_documento = %s
                        """
                        DatabaseManager.execute_query(update_pdf_query, (pdf_content, id_documento))
                        
                        self.mostrar_mensaje(
                            "Éxito",
                            "PDF actualizado correctamente",
                            QMessageBox.Icon.Information
                        )
                except Exception as e:
                    self.mostrar_mensaje(
                        "Error",
                        f"Error al actualizar PDF: {str(e)}",
                        QMessageBox.Icon.Critical
                    )

            pdf_button.clicked.connect(cambiar_pdf)
            form_layout.addRow("PDF:", pdf_button)

            layout.addLayout(form_layout)

            # Botones
            button_box = QDialogButtonBox(
                QDialogButtonBox.StandardButton.Save | 
                QDialogButtonBox.StandardButton.Cancel
            )

            def guardar_modificacion():
                try:
                    # Construir la consulta de actualización
                    query = """
                        UPDATE documento 
                        SET fecha = %s, establecimiento = %s, tipodocumento = %s,
                            nrodocumento = %s, materia = %s,
                            firma = %s, estado = %s
                        WHERE id_documento = %s
                    """
                    
                    valores = [
                        fecha_input.selectedDate().toString("yyyy-MM-dd"),
                        inputs['establecimiento'].currentText(),
                        inputs['tipodocumento'].currentText(),
                        inputs['nrodocumento'].text(),
                        inputs['materia'].text(),
                        inputs['firma'].text(),
                        inputs['estado'].text(),
                        id_documento
                    ]

                    DatabaseManager.execute_query(query, valores)

                    self.mostrar_mensaje(
                        "Éxito",
                        "Documento modificado correctamente",
                        QMessageBox.Icon.Information
                    )

                    dialog.accept()
                    self.consultar_datos()

                except Exception as e:
                    self.mostrar_mensaje(
                        "Error",
                        f"Error al modificar documento: {str(e)}",
                        QMessageBox.Icon.Critical
                    )

            button_box.accepted.connect(guardar_modificacion)
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)

            # Aplicar estilos
            dialog.setStyleSheet(f"""
                QDialog {{
                    background-color: {COLORS['background']};
                }}
                QLabel {{
                    color: {COLORS['text']};
                }}
                QLineEdit, QComboBox {{
                    background-color: {COLORS['surface']};
                    color: {COLORS['text']};
                    padding: 8px;
                    border: 1px solid {COLORS['primary']};
                    border-radius: 4px;
                }}
            """)

            dialog.exec()

        except Exception as e:
            self.mostrar_mensaje(
                "Error",
                f"Error al abrir formulario de modificación: {str(e)}",
                QMessageBox.Icon.Critical
            )

    def mostrar_mensaje(self, titulo, mensaje, icono=QMessageBox.Icon.Information):
        """Muestra un mensaje en una ventana de diálogo"""
        msg = QMessageBox(self)
        msg.setWindowTitle(titulo)
        msg.setText(mensaje)
        msg.setIcon(icono)
        msg.setStyleSheet(f"""
            QMessageBox {{
                background-color: {COLORS['background']};
            }}
            QMessageBox QLabel {{
                color: {COLORS['text']};
            }}
            QPushButton {{
                background-color: {COLORS['primary']};
                color: white;
                padding: 6px 20px;
                border: none;
                border-radius: 4px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['primary_light']};
            }}
        """)
        return msg.exec()

    def enviar_documento(self, documento):
        try:
            # Verificar que el documento esté en el departamento actual
            if documento['lugar_actual'] != self.departamento:
                self.mostrar_mensaje(
                    "Error",
                    "Solo puedes enviar documentos que estén en tu departamento actual",
                    QMessageBox.Icon.Warning
                )
                return

            # Crear diálogo de confirmación
            dialog = QDialog(self)
            dialog.setWindowTitle("Enviar Documento")
            dialog.setFixedWidth(500)
            layout = QVBoxLayout(dialog)

            # Título
            title_label = QLabel("Información del Documento")
            title_label.setStyleSheet(f"""
                QLabel {{
                    color: {COLORS['text']};
                    font-size: 16px;
                    font-weight: bold;
                    padding: 10px;
                }}
            """)
            layout.addWidget(title_label)

            # Crear un widget para la información detallada
            info_widget = QWidget()
            info_layout = QFormLayout(info_widget)
            info_widget.setStyleSheet(f"""
                QWidget {{
                    background-color: {COLORS['surface']};
                    border-radius: 8px;
                    padding: 15px;
                }}
                QLabel {{
                    color: {COLORS['text']};
                    padding: 5px;
                }}
            """)

            # Campos de información
            campos = [
                ("ID:", str(documento['id_documento'])),
                ("Tipo de Documento:", documento['tipodocumento']),
                ("Número:", documento['nrodocumento']),
                ("Materia:", documento.get('materia', 'No especificada')),
                ("Establecimiento:", documento.get('establecimiento', 'No especificado')),
                ("Ubicación Actual:", documento['lugar_actual']),
                ("Firma:", documento.get('firma', 'No especificada')),
                ("Estado:", documento.get('estado', 'No especificado'))
            ]

            for label, valor in campos:
                label_widget = QLabel(label)
                valor_widget = QLabel(valor)
                valor_widget.setStyleSheet(f"color: {COLORS['text_secondary']};")
                info_layout.addRow(label_widget, valor_widget)

            layout.addWidget(info_widget)

            # Selector de departamento destino
            destino_widget = QWidget()
            destino_layout = QFormLayout(destino_widget)
            destino_widget.setStyleSheet(f"""
                QWidget {{
                    background-color: {COLORS['surface']};
                    border-radius: 8px;
                    padding: 15px;
                    margin-top: 10px;
                }}
            """)

            # Obtener lista actualizada de departamentos
            departamentos = DatabaseManager.get_departamentos()
            departamento_combo = QComboBox()
            if departamentos:  # Verificar que la lista no esté vacía
                departamento_combo.addItems(departamentos)
                # Establecer el departamento actual como valor predeterminado si existe
                if 'destino' in documento and documento['destino'] in departamentos:
                    departamento_combo.setCurrentText(documento['destino'])
            else:
                self.mostrar_mensaje(
                    "Error",
                    "No se pudieron cargar los departamentos",
                    QMessageBox.Icon.Warning
                )
                return

            departamento_combo.setStyleSheet(f"""
                QComboBox {{
                    background-color: {COLORS['background']};
                    color: {COLORS['text']};
                    padding: 8px;
                    border: 1px solid {COLORS['primary']};
                    border-radius: 4px;
                }}
                QComboBox:hover {{
                    border: 1px solid {COLORS['primary_light']};
                }}
                QComboBox::drop-down {{
                    border: none;
                }}
                QComboBox::down-arrow {{
                    image: url(down_arrow.png);
                    width: 12px;
                    height: 12px;
                }}
                QComboBox QAbstractItemView {{
                    background-color: {COLORS['background']};
                    color: {COLORS['text']};
                    selection-background-color: {COLORS['primary']};
                    selection-color: white;
                    border: 1px solid {COLORS['primary']};
                }}
            """)

            destino_layout.addRow("Departamento Destino:", departamento_combo)
            layout.addWidget(destino_widget)

            # Botones
            button_container = QWidget()
            button_layout = QHBoxLayout(button_container)
            
            # Botón Enviar
            enviar_btn = QPushButton("Enviar")
            enviar_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {COLORS['primary']};
                    color: {COLORS['text']};
                    border: none;
                    padding: 8px 16px;
                    border-radius: 6px;
                    font-weight: bold;
                    min-width: 100px;
                }}
                QPushButton:hover {{
                    background-color: {COLORS['primary_light']};
                }}
            """)
            enviar_btn.clicked.connect(lambda: self.confirmar_envio(documento, departamento_combo, dialog))  # Pasamos el combo box completo
            
            # Botón Cancelar
            cancelar_btn = QPushButton("Cancelar")
            cancelar_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {COLORS['surface']};
                    color: {COLORS['text']};
                    border: 2px solid {COLORS['error']};
                    padding: 8px 16px;
                    border-radius: 6px;
                    font-weight: bold;
                    min-width: 100px;
                }}
                QPushButton:hover {{
                    background-color: {COLORS['error']};
                }}
            """)
            cancelar_btn.clicked.connect(dialog.reject)
            
            button_layout.addWidget(enviar_btn)
            button_layout.addWidget(cancelar_btn)
            layout.addWidget(button_container)

            dialog.setStyleSheet(f"""
                QDialog {{
                    background-color: {COLORS['background']};
                }}
            """)
            
            dialog.exec()

        except Exception as e:
            self.mostrar_mensaje("Error", f"Error al enviar documento: {str(e)}")

    def confirmar_envio(self, documento, departamento_combo, dialog):
        try:
            # Obtener el departamento seleccionado del combo box
            departamento_destino = departamento_combo.currentText()
            
            # Primero verificamos si existe una solicitud pendiente para este documento
            check_query = """
                SELECT id_solicitud 
                FROM solicitudes_documento 
                WHERE id_documento = %s 
                AND estado = 'pendiente'
            """
            solicitud_existente = DatabaseManager.execute_query(check_query, (documento['id_documento'],))

            if solicitud_existente:
                # Si existe una solicitud pendiente, actualizamos el destino
                update_query = """
                    UPDATE solicitudes_documento 
                    SET departamento_destino = %s,
                        fecha_solicitud = CONVERT_TZ(NOW(), 'UTC', 'America/Santiago')
                    WHERE id_documento = %s 
                    AND estado = 'pendiente'
                """
                values = (departamento_destino, documento['id_documento'])
                DatabaseManager.execute_query(update_query, values)
                
                # Actualizar el destino en la tabla documento
                update_doc_query = """
                    UPDATE documento 
                    SET destino = %s 
                    WHERE id_documento = %s
                """
                DatabaseManager.execute_query(update_doc_query, values)
                
                mensaje = "Solicitud de envío actualizada correctamente"
            else:
                # Si no existe, creamos una nueva solicitud
                insert_query = """
                    INSERT INTO solicitudes_documento 
                    (id_documento, departamento_origen, departamento_destino, fecha_solicitud) 
                    VALUES (%s, %s, %s, CONVERT_TZ(NOW(), 'UTC', 'America/Santiago'))
                """
                values = (
                    documento['id_documento'],
                    documento['lugar_actual'],
                    departamento_destino
                )
                DatabaseManager.execute_query(insert_query, values)
                
                # Actualizar el destino en la tabla documento
                update_doc_query = """
                    UPDATE documento 
                    SET destino = %s 
                    WHERE id_documento = %s
                """
                doc_values = (departamento_destino, documento['id_documento'])
                DatabaseManager.execute_query(update_doc_query, doc_values)
                
                mensaje = "Solicitud de envío creada correctamente"
            
            self.mostrar_mensaje(
                "Éxito",
                mensaje,
                QMessageBox.Icon.Information
            )
            
            dialog.accept()
            self.consultar_datos()  # Actualizar vista
            
        except Exception as e:
            self.mostrar_mensaje(
                "Error",
                f"Error al crear solicitud: {str(e)}",
                QMessageBox.Icon.Critical
            )

    def recibir_documento(self):
        try:
            # Cerrar conexión existente si hay una
            if hasattr(DatabaseManager, '_connection_pool') and DatabaseManager._connection_pool:
                DatabaseManager._connection_pool.close()
                DatabaseManager._connection_pool = None
                
            # Obtener nueva conexión
            connection = DatabaseManager.get_connection()
            
            query = """
                SELECT s.*, d.tipodocumento, d.nrodocumento, d.materia, 
                       d.establecimiento, d.firma, d.estado
                FROM solicitudes_documento s
                JOIN documento d ON s.id_documento = d.id_documento
                WHERE s.departamento_destino = %s 
                AND s.estado = 'pendiente'
                ORDER BY s.fecha_solicitud DESC
            """
            solicitudes = DatabaseManager.execute_query(query, (self.departamento,))
            
            if not solicitudes:
                self.mostrar_mensaje(
                    "Información",
                    "No hay documentos pendientes por recibir",
                    QMessageBox.Icon.Information
                )
                return

            dialog = QDialog(self)
            dialog.setWindowTitle("Documentos Pendientes de Recepción")
            dialog.setFixedWidth(800)
            layout = QVBoxLayout(dialog)

            tabla = QTableWidget()
            tabla.setColumnCount(8)
            tabla.setHorizontalHeaderLabels([
                "ID", "Tipo Doc", "Nro Doc", "Materia", 
                "Origen", "Fecha Solicitud", "Estado", "Acciones"
            ])
            tabla.setStyleSheet(f"""
                QTableWidget {{
                    background-color: {COLORS['surface']};
                    color: {COLORS['text']};
                    border: none;
                    border-radius: 8px;
                }}
                QTableWidget::item {{
                    padding: 5px;
                }}
                QHeaderView::section {{
                    background-color: {COLORS['primary']};
                    color: {COLORS['text']};
                    padding: 8px;
                    border: none;
                }}
            """)

            tabla.setRowCount(len(solicitudes))
            for i, solicitud in enumerate(solicitudes):
                tabla.setItem(i, 0, QTableWidgetItem(str(solicitud['id_documento'])))
                tabla.setItem(i, 1, QTableWidgetItem(solicitud['tipodocumento']))
                tabla.setItem(i, 2, QTableWidgetItem(solicitud['nrodocumento']))
                tabla.setItem(i, 3, QTableWidgetItem(solicitud['materia']))
                tabla.setItem(i, 4, QTableWidgetItem(solicitud['departamento_origen']))
                tabla.setItem(i, 5, QTableWidgetItem(str(solicitud['fecha_solicitud'])))
                tabla.setItem(i, 6, QTableWidgetItem(solicitud['estado']))

                action_widget = QWidget()
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(2, 2, 2, 2)
                action_layout.setSpacing(4)

                aceptar_btn = QPushButton("✓")
                aceptar_btn.setToolTip("Aceptar documento")
                aceptar_btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {COLORS['success']};
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 5px;
                        font-size: 16px;
                    }}
                    QPushButton:hover {{
                        background-color: #43A047;
                    }}
                """)
                aceptar_btn.clicked.connect(
                    lambda _, s=solicitud: self.procesar_recepcion(s, True, dialog))

                rechazar_btn = QPushButton("✗")
                rechazar_btn.setToolTip("Rechazar documento")
                rechazar_btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {COLORS['error']};
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 5px;
                        font-size: 16px;
                    }}
                    QPushButton:hover {{
                        background-color: #E53935;
                    }}
                """)
                rechazar_btn.clicked.connect(
                    lambda _, s=solicitud: self.procesar_recepcion(s, False, dialog)
                )

                action_layout.addWidget(aceptar_btn)
                action_layout.addWidget(rechazar_btn)
                tabla.setCellWidget(i, 7, action_widget)

            tabla.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            tabla.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
            
            layout.addWidget(tabla)
            
            cerrar_btn = QPushButton("Cerrar")
            cerrar_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {COLORS['primary']};
                    color: white;
                    border: none;
                    padding: 8px 16px;
                    border-radius: 6px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: {COLORS['primary_dark']};
                }}
            """)
            cerrar_btn.clicked.connect(dialog.close)
            layout.addWidget(cerrar_btn, alignment=Qt.AlignmentFlag.AlignCenter)

            dialog.setStyleSheet(f"""
                QDialog {{
                    background-color: {COLORS['background']};
                }}
            """)
            
            dialog.exec()

        except Exception as e:
            self.mostrar_mensaje(
                "Error",
                f"Error al cargar documentos pendientes: {str(e)}",
                QMessageBox.Icon.Critical
            )

    def procesar_recepcion(self, solicitud, aceptar, dialog):
        try:
            estado = 'aceptado' if aceptar else 'rechazado'
            motivo = None

            if not aceptar:
                motivo, ok = QInputDialog.getText(
                    self, 
                    "Motivo de Rechazo",
                    "Por favor, indique el motivo del rechazo:",
                    QLineEdit.EchoMode.Normal
                )
                if not ok or not motivo.strip():
                    return

            # Actualizar la solicitud
            update_solicitud = """
                UPDATE solicitudes_documento 
                SET estado = %s,
                    fecha_respuesta = CONVERT_TZ(NOW(), 'UTC', 'America/Santiago'),
                    motivo_rechazo = %s
                WHERE id_documento = %s 
                AND estado = 'pendiente'
            """
            DatabaseManager.execute_query(
                update_solicitud, 
                (estado, motivo, solicitud['id_documento'])
            )

            if aceptar:
                # Actualizar documento: lugar_actual = nuevo departamento y destino = vacío
                update_documento = """
                    UPDATE documento 
                    SET lugar_actual = %s,
                        destino = ''
                    WHERE id_documento = %s
                """
                DatabaseManager.execute_query(
                    update_documento, 
                    (self.departamento, solicitud['id_documento'])
                )

                mensaje = "Documento recibido correctamente"
            else:
                # Si se rechaza, mantener lugar_actual y limpiar destino
                update_documento = """
                    UPDATE documento 
                    SET destino = ''
                    WHERE id_documento = %s
                """
                DatabaseManager.execute_query(
                    update_documento, 
                    (solicitud['id_documento'],)
                )

                mensaje = "Documento rechazado correctamente"

            self.mostrar_mensaje(
                "Éxito",
                mensaje,
                QMessageBox.Icon.Information
            )
            
            # Actualizar la vista y cerrar el diálogo
            self.consultar_datos()
            dialog.accept()

        except Exception as e:
            self.mostrar_mensaje(
                "Error",
                f"Error al procesar la recepción: {str(e)}",
                QMessageBox.Icon.Critical
            )

    def generar_reporte_actual(self):
        try:
            # Crear un nuevo libro de trabajo
            wb = Workbook()
            
            # Crear hoja de resumen
            ws_resumen = wb.active
            ws_resumen.title = "Resumen"
            
            # Obtener conteos
            query_conteos = """
                SELECT 
                    SUM(CASE WHEN estado = 'Pendiente' THEN 1 ELSE 0 END) as pendientes,
                    SUM(CASE WHEN estado = 'Finalizado' THEN 1 ELSE 0 END) as finalizados,
                    SUM(CASE WHEN estado = 'Recepcionado' THEN 1 ELSE 0 END) as recepcionados,
                    COUNT(*) as total
                FROM documento
            """
            resultados_conteo = DatabaseManager.execute_query(query_conteos)[0]
            
            # Crear tabla de resumen
            ws_resumen['A1'] = "Resumen de Documentos"
            ws_resumen['A1'].font = Font(bold=True, size=14)
            ws_resumen.merge_cells('A1:B1')
            
            # Agregar datos del resumen
            headers_resumen = [
                ("Documentos Pendientes", resultados_conteo['pendientes'] or 0),
                ("Documentos Finalizados", resultados_conteo['finalizados'] or 0),
                ("Documentos Recepcionados", resultados_conteo['recepcionados'] or 0),
                ("Total de Documentos", resultados_conteo['total'] or 0)
            ]
            
            for idx, (header, value) in enumerate(headers_resumen, start=3):
                ws_resumen[f'A{idx}'] = header
                ws_resumen[f'B{idx}'] = value
                ws_resumen[f'A{idx}'].font = Font(bold=True)
            
            # Ajustar ancho de columnas en resumen
            ws_resumen.column_dimensions['A'].width = 25
            ws_resumen.column_dimensions['B'].width = 15
            
            # Agregar estilos a la tabla de resumen
            for row in range(3, len(headers_resumen) + 3):
                for col in ['A', 'B']:
                    cell = ws_resumen[f'{col}{row}']
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # Crear hoja de detalles
            ws_detalles = wb.create_sheet("Detalles")
            
            # Definir encabezados para la hoja de detalles
            headers = [
                "ID", "Fecha", "Establecimiento", "Tipo Doc", 
                "Nro Doc", "Materia", "Lugar Actual", "Destino", 
                "Firma", "Estado"
            ]
            
            # Agregar encabezados a la hoja de detalles
            for col, header in enumerate(headers, 1):
                cell = ws_detalles.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # Obtener datos de la tabla y agregarlos a la hoja de detalles
            for row in range(self.tree_widget.topLevelItemCount()):
                item = self.tree_widget.topLevelItem(row)
                if not item.isHidden():  # Solo exportar elementos visibles
                    # Mapear las columnas en el orden especificado
                    valores = [
                        item.text(0),  # ID
                        item.text(1),  # Fecha
                        item.text(2),  # Establecimiento
                        item.text(3),  # Tipo Doc
                        item.text(4),  # Nro Doc
                        item.text(5),  # Materia
                        item.text(6),  # Lugar Actual
                        item.text(7),  # Destino
                        item.text(8),  # Firma
                        item.text(9),  # Estado
                    ]
                    
                    for col, valor in enumerate(valores, 1):
                        cell = ws_detalles.cell(row=row+2, column=col, value=valor)
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

            # Ajustar el ancho de las columnas en la hoja de detalles
            for column in ws_detalles.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_detalles.column_dimensions[column_letter].width = adjusted_width

            # Obtener la fecha actual para el nombre del archivo
            fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Abrir diálogo para guardar archivo
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Reporte",
                f"Reporte_Documentos_{fecha_actual}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if file_name:
                # Asegurarse de que el archivo termine en .xlsx
                if not file_name.endswith('.xlsx'):
                    file_name += '.xlsx'
                    
                # Guardar el archivo
                wb.save(file_name)
                
                self.mostrar_mensaje(
                    "Éxito",
                    f"Reporte generado exitosamente:\n{file_name}",
                    QMessageBox.Icon.Information
                )

        except Exception as e:
            self.mostrar_mensaje(
                "Error",
                f"Error al generar el reporte: {str(e)}",
                QMessageBox.Icon.Critical
            )

    def descargar_pdf(self, doc_id):
        try:
            # Obtener el PDF de la base de datos
            query = "SELECT archivo_pdf FROM documento WHERE id_documento = %s"
            resultado = DatabaseManager.execute_query(query, (doc_id,))
            
            if not resultado or not resultado[0]['archivo_pdf']:
                self.mostrar_mensaje(
                    "Error",
                    "No se encontró el archivo PDF para este documento",
                    QMessageBox.Icon.Warning
                )
                return

            # Abrir diálogo para guardar archivo
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar PDF",
                f"documento_{doc_id}.pdf",
                "PDF Files (*.pdf)"
            )

            if file_name:
                # Asegurarse de que el archivo termine en .pdf
                if not file_name.endswith('.pdf'):
                    file_name += '.pdf'
                
                # Escribir el PDF al archivo
                with open(file_name, 'wb') as f:
                    f.write(resultado[0]['archivo_pdf'])
                
                self.mostrar_mensaje(
                    "Éxito",
                    f"PDF guardado exitosamente en:\n{file_name}",
                    QMessageBox.Icon.Information
                )

        except Exception as e:
            self.mostrar_mensaje(
                "Error",
                f"Error al descargar el PDF: {str(e)}",
                QMessageBox.Icon.Critical
            )
            if hasattr(DatabaseManager, '_connection_pool') and DatabaseManager._connection_pool:
                DatabaseManager._connection_pool.close()
                DatabaseManager._connection_pool = None

import json
import os

class CredentialManager:
    CACHE_FILE = 'credentials_cache.json'
    
    @classmethod
    def save_credentials(cls, email, password, remember=True):
        """Guarda las credenciales del usuario"""
        try:
            if not remember:
                cls.clear_credentials()
                return
                
            data = {
                'email': email,  # Cambiado de username a email
                'password': password
            }
            
            with open(cls.CACHE_FILE, 'w') as f:
                json.dump(data, f)
                
        except Exception as e:
            print(f"Error al guardar credenciales: {e}")

    @classmethod
    def load_credentials(cls):
        """Carga las credenciales guardadas"""
        try:
            if not os.path.exists(cls.CACHE_FILE):
                return None
                
            with open(cls.CACHE_FILE, 'r') as f:
                return json.load(f)
                
        except Exception as e:
            print(f"Error al cargar credenciales: {e}")
            return None

    @classmethod
    def clear_credentials(cls):
        """Elimina el archivo de credenciales"""
        try:
            if os.path.exists(cls.CACHE_FILE):
                os.remove(cls.CACHE_FILE)
        except Exception as e:
            print(f"Error al eliminar credenciales: {e}")

class LoginDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Inicio de Sesión - Corporación Isla de Maipo")
        self.setFixedWidth(550)
        self.setFixedHeight(650)
        self.user_role = None
        self.user_departamento = None  # Nuevo atributo para el departamento
        self.setup_ui()
        
        # Cargar credenciales guardadas
        saved_credentials = CredentialManager.load_credentials()
        if saved_credentials:
            self.email_input.setText(saved_credentials.get('email', ''))  # Cambiado de username a email
            self.password_input.setText(saved_credentials.get('password', ''))
            self.remember_checkbox.setChecked(True)

    def setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(40, 30, 40, 30)

        # Logo
        logo_label = QLabel()
        logo_pixmap = QPixmap(resource_path("isla_de_maipo.png"))
        scaled_pixmap = logo_pixmap.scaled(360, 360, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
        logo_label.setPixmap(scaled_pixmap)
        logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(logo_label)

        # Ttulo
        title_label = QLabel("Bienvenido")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet("""
            QLabel {
                color: #ffffff;
                font-size: 24px;
                font-weight: bold;
                margin-bottom: 20px;
            }
        """)
        main_layout.addWidget(title_label)

        # Formulario
        form_widget = QWidget()
        form_layout = QFormLayout(form_widget)
        form_layout.setSpacing(15)

        # Modificar el campo de usuario por email
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("Ingrese su email")
        self.email_input.setStyleSheet(create_input_style())
        self.email_input.setMinimumHeight(42)
        self.email_input.setFont(QFont("Segoe UI", 14))
        self.email_input.returnPressed.connect(self.login)

        # Contenedor para el campo de contraseña
        password_container = QWidget()
        password_layout = QHBoxLayout(password_container)
        password_layout.setContentsMargins(0, 0, 0, 0)
        password_layout.setSpacing(5)
        
        self.password_input = QLineEdit()
        self.password_input.setPlaceholderText("Ingrese su contraseña")
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_input.setStyleSheet(create_input_style())
        self.password_input.setMinimumHeight(42)
        self.password_input.setFont(QFont("Segoe UI", 14))
        self.password_input.returnPressed.connect(self.login)
        password_layout.addWidget(self.password_input)
        
        # Botón de visibilidad para contraseña
        self.toggle_password_btn = QPushButton("🔒")
        self.toggle_password_btn.setFixedSize(35, 35)
        self.toggle_password_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.toggle_password_btn.setStyleSheet(self.create_visibility_button_style())
        self.toggle_password_btn.clicked.connect(self.toggle_password_visibility)
        password_layout.addWidget(self.toggle_password_btn)

        # Agregar los contenedores al formulario
        form_layout.addRow(self.create_label("Email:"), self.email_input)
        form_layout.addRow(self.create_label("Contraseña:"), password_container)

        main_layout.addWidget(form_widget)
    
        # Agregar checkbox "Recérdame" antes de los botones
        self.remember_checkbox = QCheckBox("Recordar mis datos")
        self.remember_checkbox.setStyleSheet(f"""
            QCheckBox {{
                color: {COLORS['text']};
                font-size: 13px;
                spacing: 8px;
            }}
            QCheckBox::indicator {{
                width: 18px;
                height: 18px;
                border: 2px solid {COLORS['primary']};
                border-radius: 4px;
                background-color: {COLORS['surface']};
            }}
            QCheckBox::indicator:checked {{
                background-color: {COLORS['primary']};
                image: url("check.png");
            }}
            QCheckBox::indicator:hover {{
                border-color: {COLORS['primary_light']};
            }}
        """)
        main_layout.addWidget(self.remember_checkbox)

        # Botones
        buttons_layout = QHBoxLayout()  # Cambiado a QHBoxLayout
        buttons_layout.setSpacing(20)  # Espaciado entre botones
        buttons_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)  # Centrar botones

        # Botón de inicio de sesión
        login_btn = QPushButton("Iniciar Sesión")
        login_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['primary']};
                color: {COLORS['text']};
                border: none;
                padding: 12px;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: {COLORS['primary_dark']};
            }}
            QPushButton:pressed {{
                background-color: {COLORS['primary']};
            }}
        """)
        login_btn.clicked.connect(self.login)
        buttons_layout.addWidget(login_btn)

        # Botón de registro
        register_btn = QPushButton("Registrarse")
        register_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: {COLORS['primary']};
                border: 2px solid {COLORS['primary']};
                padding: 12px;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: rgba(33, 150, 243, 0.1);
            }}
            QPushButton:pressed {{
                background-color: rgba(33, 150, 243, 0.2);
            }}
        """)
        register_btn.clicked.connect(self.show_register)
        buttons_layout.addWidget(register_btn)

        main_layout.addLayout(buttons_layout)

        # Estilo general del diálogo
        self.setStyleSheet("""
            QDialog {
                background-color: #2b2b2b;
            }
        """)

    def create_label(self, text):
        label = QLabel(text)
        label.setStyleSheet("""
            QLabel {
                color: #ffffff;
                font-size: 14px;
                font-weight: bold;
            }
        """)
        return label

    def login(self):
        try:
            email = self.email_input.text()
            password = self.password_input.text()
            
            if not email or not password:
                self.show_custom_error(
                    "Campos Incompletos", 
                    "Por favor complete todos los campos para iniciar sesión.",
                    "Los campos de email y contraseña son obligatorios."
                )
                return

            # Validar formato de email básico
            if '@' not in email or '.' not in email:
                self.show_custom_error(
                    "Email Inválido",
                    "Por favor ingrese un email válido.",
                    "El formato debe ser ejemplo@dominio.com"
                )
                return
                
            success, role, departamento = DatabaseManager.validate_login(email, password)
            if success:
                self.user_role = role
                self.user_departamento = departamento  # Guardar el departamento
                
                # Guardar credenciales si el checkbox está marcado
                if self.remember_checkbox.isChecked():
                    CredentialManager.save_credentials(
                        email, 
                        password, 
                        remember=True
                    )
                else:
                    CredentialManager.clear_credentials()
                
                self.accept()
            else:
                self.show_custom_error(
                    "Error de Autenticación", 
                    "No se pudo iniciar sesión con las credenciales proporcionadas.",
                    "Por favor verifique su email y contraseña."
                )
        except Exception as e:
            self.show_error_message(str(e))

    def show_custom_error(self, title, message, detail):
        error_dialog = QDialog(self)
        error_dialog.setWindowTitle(title)
        error_dialog.setFixedWidth(400)
        
        layout = QVBoxLayout(error_dialog)
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)

        # Icono de error
        icon_label = QLabel()
        icon_label.setText("⚠️")
        icon_label.setStyleSheet("""
            QLabel {
                color: #EF5350;
                font-size: 48px;
            }
        """)
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(icon_label)

        # Mensaje principal
        message_label = QLabel(message)
        message_label.setWordWrap(True)
        message_label.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text']};
                font-size: 16px;
                font-weight: bold;
            }}
        """)
        message_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(message_label)

        # Detalle
        detail_label = QLabel(detail)
        detail_label.setWordWrap(True)
        detail_label.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text_secondary']};
                font-size: 14px;
            }}
        """)
        detail_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(detail_label)

        # Botón de cerrar
        close_btn = QPushButton("Entendido")
        close_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['primary']};
                color: {COLORS['text']};
                border: none;
                padding: 12px;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
                min-width: 100px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['primary_dark']};
            }}
            QPushButton:pressed {{
                background-color: {COLORS['primary']};
            }}
        """)
        close_btn.clicked.connect(error_dialog.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignmentFlag.AlignCenter)

        # Estilo general del diálogo
        error_dialog.setStyleSheet("""
            QDialog {
                background-color: {COLORS['background']};
            }
        """)

        error_dialog.exec()

    def show_register(self):
        dialog = RegisterDialog(self)
        dialog.exec()

    def get_user_role(self):
        return self.user_role

    def get_user_departamento(self):  # Nuevo método
        """Retorna el departamento del usuario"""
        return self.user_departamento

    def toggle_password_visibility(self):
        if self.password_input.echoMode() == QLineEdit.EchoMode.Password:
            self.password_input.setEchoMode(QLineEdit.EchoMode.Normal)
            self.toggle_password_btn.setText("🔓")
        else:
            self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
            self.toggle_password_btn.setText("🔒")

    def toggle_confirm_visibility(self):
        if self.confirm_password_input.echoMode() == QLineEdit.EchoMode.Password:
            self.confirm_password_input.setEchoMode(QLineEdit.EchoMode.Normal)
            self.toggle_confirm_btn.setText("🔓")
        else:
            self.confirm_password_input.setEchoMode(QLineEdit.EchoMode.Password)
            self.toggle_confirm_btn.setText("🔒")

    def create_visibility_button_style(self):
        return f"""
            QPushButton {{
                background-color: {COLORS['surface']};
                border: none;
                border-radius: 4px;
                padding: 5px;
                font-size: 18px;
                color: {COLORS['text']};
            }}
            QPushButton:hover {{
                background-color: {COLORS['primary']};
            }}
        """

    def get_user_email(self):  # Nuevo método para obtener el email
        """Retorna el email del usuario"""
        return self.email_input.text()

class RegisterDialog(QDialog):
    def __init__(self, parent=None, admin_mode=False):
        super().__init__(parent)
        self.setWindowTitle("Registro de Usuario - Corporación Isla de Maipo")
        self.setFixedWidth(600)  # Reducido el ancho
        self.setFixedHeight(750)  # Reducido el alto
        self.admin_mode = admin_mode
        self.setup_ui()

    def setup_ui(self):
        # Layout principal con márgenes ajustados
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)  # Reducido el espaciado
        main_layout.setContentsMargins(40, 20, 40, 20)  # Reducido márgenes superior e inferior

        # Logo
        logo_label = QLabel()
        logo_pixmap = QPixmap(resource_path("isla_de_maipo.png"))
        scaled_pixmap = logo_pixmap.scaled(170, 170, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
        logo_label.setPixmap(scaled_pixmap)
        logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_layout.addWidget(logo_label)

        # Título con estilo mejorado
        title_label = QLabel("Registro de Usuario")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text']};
                font-size: 24px;
                font-weight: bold;
                margin: 20px 0;
                padding: 15px;
                background-color: {COLORS['surface']};
                border-radius: 8px;
            }}
        """)
        main_layout.addWidget(title_label)

        # Formulario
        form_widget = QWidget()
        form_layout = QFormLayout(form_widget)
        form_layout.setSpacing(20)  # Aumentado el espaciado entre campos
        form_layout.setContentsMargins(10, 10, 10, 10)  # Márgenes internos

        # Campos de entrada con estilos mejorados
        self.username_input = self.create_input("Ingrese un nombre de usuario")
        self.email_input = self.create_input("Ingrese su correo electrónico")
        
        # Contenedor para contraseña
        password_container = QWidget()
        password_layout = QHBoxLayout(password_container)
        password_layout.setContentsMargins(0, 0, 0, 0)
        password_layout.setSpacing(10)
        
        self.password_input = self.create_input("Ingrese una contraseña segura", is_password=True)
        password_layout.addWidget(self.password_input)
        
        # Botón de visibilidad para contraseña
        self.toggle_password_btn = QPushButton("🔒")
        self.toggle_password_btn.setFixedSize(42, 42)  # Botón más grande
        self.toggle_password_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.toggle_password_btn.setStyleSheet(self.create_visibility_button_style())
        self.toggle_password_btn.clicked.connect(self.toggle_password_visibility)
        password_layout.addWidget(self.toggle_password_btn)

        # Contenedor para confirmar contraseña
        confirm_container = QWidget()
        confirm_layout = QHBoxLayout(confirm_container)
        confirm_layout.setContentsMargins(0, 0, 0, 0)
        confirm_layout.setSpacing(10)
        
        self.confirm_password_input = self.create_input("Confirme su contraseña", is_password=True)
        confirm_layout.addWidget(self.confirm_password_input)
        
        # Botón de visibilidad para confirmar contraseña
        self.toggle_confirm_btn = QPushButton("🔒")
        self.toggle_confirm_btn.setFixedSize(42, 42)
        self.toggle_confirm_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.toggle_confirm_btn.setStyleSheet(self.create_visibility_button_style())
        self.toggle_confirm_btn.clicked.connect(self.toggle_confirm_visibility)
        confirm_layout.addWidget(self.toggle_confirm_btn)

        # Agregar ComboBox para departamento
        self.departamento_combo = QComboBox()
        self.departamento_combo.setStyleSheet(f"""
            QComboBox {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                padding: 12px 15px;
                border: 2px solid {COLORS['surface']};
                border-radius: 6px;
                min-width: 350px;
                font-size: 14px;
            }}
            QComboBox:hover {{
                border: 2px solid {COLORS['primary']};
            }}
            QComboBox::drop-down {{
                border: none;
                padding-right: 20px;
            }}
            QComboBox::down-arrow {{
                image: url(down_arrow.png);
                width: 12px;
                height: 12px;
            }}
            QComboBox QAbstractItemView {{
                background-color: {COLORS['background']};
                color: {COLORS['text']};
                selection-background-color: {COLORS['primary']};
                selection-color: white;
            }}
        """)
        
        # Obtener y cargar departamentos
        departamentos = DatabaseManager.get_departamentos()
        self.departamento_combo.addItems(departamentos)

        # Agregar los campos al formulario con espaciado
        form_layout.addRow(self.create_label("Usuario:"), self.username_input)
        form_layout.addRow(self.create_label("Email:"), self.email_input)
        form_layout.addRow(self.create_label("Contraseña:"), password_container)
        form_layout.addRow(self.create_label("Confirmar:"), confirm_container)
        form_layout.addRow(self.create_label("Departamento:"), self.departamento_combo)

        # Agregar ComboBox para rol si es modo admin
        if self.admin_mode:
            self.role_combo = QComboBox()
            self.role_combo.addItems(["usuario", "recepcionista", "admin"])
            self.role_combo.setStyleSheet(self.create_combo_style())
            self.role_combo.setFixedHeight(42)  # Altura consistente
            form_layout.addRow(self.create_label("Rol:"), self.role_combo)

        main_layout.addWidget(form_widget)

        # Ajustar espaciado antes de los botones (reducido para subirlos)
        main_layout.addSpacing(10)  # Reducido el espaciado antes de los botones

            # Botones
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(15)
        buttons_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Botón de registro
        register_btn = QPushButton("Registrar Usuario")
        register_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        register_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['primary']};
                color: {COLORS['text']};
                border: none;
                padding: 8px 16px;
                border-radius: 6px;
                font-size: 13px;
                font-weight: bold;
                min-width: 140px;
                height: 35px;
                text-transform: uppercase;
                letter-spacing: 0.5px;
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
                transition: all 0.3s ease;
            }}
            QPushButton:hover {{
                background-color: {COLORS['primary_light']};
                transform: translateY(-2px);
                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.2);
            }}
            QPushButton:pressed {{
                background-color: {COLORS['primary_dark']};
                transform: translateY(1px);
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
            }}
        """)
        register_btn.clicked.connect(self.register)
        buttons_layout.addWidget(register_btn)

        # Botón cancelar
        cancel_btn = QPushButton("Cancelar")
        cancel_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        cancel_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: {COLORS['text']};
                border: 2px solid {COLORS['error']};
                padding: 8px 16px;
                border-radius: 6px;
                font-size: 13px;
                font-weight: bold;
                min-width: 140px;
                height: 35px;
                text-transform: uppercase;
                letter-spacing: 0.5px;
                transition: all 0.3s ease;
            }}
            QPushButton:hover {{
                background-color: {COLORS['error']};
                color: {COLORS['text']};
                border-color: transparent;
                transform: translateY(-2px);
                box-shadow: 0 4px 6px rgba(239, 83, 80, 0.3);
            }}
            QPushButton:pressed {{
                transform: translateY(1px);
                box-shadow: 0 2px 4px rgba(239, 83, 80, 0.2);
            }}
        """)
        cancel_btn.clicked.connect(self.reject)
        buttons_layout.addWidget(cancel_btn)

        main_layout.addLayout(buttons_layout)
        
        # Reducido el espaciado después de los botones
        main_layout.addSpacing(15)

    def create_input(self, placeholder, is_password=False):
        input_field = QLineEdit()
        input_field.setPlaceholderText(placeholder)
        if is_password:
            input_field.setEchoMode(QLineEdit.EchoMode.Password)
        input_field.setStyleSheet(f"""
            QLineEdit {{
                padding: 12px 15px;
                border: 2px solid {COLORS['surface']};
                border-radius: 6px;
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                font-size: 14px;
                min-width: 350px;  # Reducido el ancho
            }}
            QLineEdit:focus {{
                border: 2px solid {COLORS['primary']};
            }}
            QLineEdit::placeholder {{
                color: {COLORS['text_secondary']};
                font-size: 13px;
                opacity: 0.7;
            }}
        """)
        input_field.setMinimumHeight(42)  # Reducido ligeramente el alto
        return input_field

    def create_label(self, text):
        label = QLabel(text)
        label.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text']};
                font-size: 14px;
                font-weight: bold;
                margin-right: 15px;
            }}
        """)
        return label

    def create_visibility_button_style(self):
        return f"""
            QPushButton {{
                background-color: {COLORS['surface']};
                border: 2px solid {COLORS['primary']};
                border-radius: 6px;
                padding: 5px;
                font-size: 18px;
                color: {COLORS['text']};
            }}
            QPushButton:hover {{
                background-color: {COLORS['primary']};
            }}
        """

    def create_combo_style(self):
        return f"""
            QComboBox {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                padding: 10px 15px;
                border: 2px solid {COLORS['primary']};
                border-radius: 6px;
                min-width: 300px;
                font-size: 14px;
            }}
            QComboBox:hover {{
                border-color: {COLORS['primary_light']};
            }}
            QComboBox::drop-down {{
                border: none;
                padding-right: 20px;
            }}
            QComboBox QAbstractItemView {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                selection-background-color: {COLORS['primary']};
                selection-color: {COLORS['text']};
                border: 1px solid {COLORS['primary']};
            }}
        """

    def register(self):
        try:
            # Obtener los valores de los campos
            username = self.username_input.text()
            email = self.email_input.text()
            password = self.password_input.text()
            rol = "usuario"  # Por defecto será usuario normal
            departamento = self.departamento_combo.currentText()

            # Validaciones
            if not all([username, email, password, departamento]):
                raise ValueError("Todos los campos son obligatorios")

            # Generar salt y hash de la contraseña
            salt = DatabaseManager.generate_salt()
            password_hash = DatabaseManager.hash_password(password, salt)

            # Registrar usuario con estado pendiente
            DatabaseManager.execute_query("""
                INSERT INTO usuario 
                (nombreusuario, email, password_hash, salt, rol, departamento, estado) 
                VALUES (%s, %s, %s, %s, %s, %s, 'pendiente')
            """, (username, email, password_hash, salt, rol, departamento))
            
            # Enviar email de notificación
            if EmailNotifier.send_registration_notification(email):
                self.show_custom_success(
                    "Registro Pendiente",
                    "¡Solicitud enviada correctamente!",
                    "Se ha enviado un correo de confirmación. Un administrador revisará su solicitud."
                )
            else:
                self.show_custom_warning(
                    "Registro Pendiente",
                    "¡Solicitud enviada correctamente!",
                    "Hubo un problema al enviar el correo de confirmación, pero su solicitud fue registrada."
                )
            
            self.accept()
            
        except Exception as e:
            self.show_custom_error(
                "Error de Registro",
                "No se pudo completar el registro.",
                str(e)
            )

    def validate_email(self, email):
        """Validación básica de formato de email"""
        import re
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, email) is not None

    def show_custom_error(self, title, message, detail):
        error_dialog = QDialog(self)
        error_dialog.setWindowTitle(title)
        error_dialog.setFixedWidth(400)
        
        layout = QVBoxLayout(error_dialog)
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)

        # Icono de error
        icon_label = QLabel()
        icon_label.setText("⚠️")
        icon_label.setStyleSheet("""
            QLabel {
                color: #EF5350;
                font-size: 48px;
            }
        """)
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(icon_label)

        # Mensaje principal
        message_label = QLabel(message)
        message_label.setWordWrap(True)
        message_label.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text']};
                font-size: 16px;
                font-weight: bold;
            }}
        """)
        message_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(message_label)

        # Detalle
        detail_label = QLabel(detail)
        detail_label.setWordWrap(True)
        detail_label.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text_secondary']};
                font-size: 14px;
            }}
        """)
        detail_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(detail_label)

        # Botón de cerrar
        close_btn = QPushButton("Entendido")
        close_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['primary']};
                color: {COLORS['text']};
                border: none;
                padding: 12px;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
                min-width: 100px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['primary_dark']};
            }}
            QPushButton:pressed {{
                background-color: {COLORS['primary']};
            }}
        """)
        close_btn.clicked.connect(error_dialog.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignmentFlag.AlignCenter)

        # Estilo general del diálogo
        error_dialog.setStyleSheet(f"""
                QDialog {{
                    background-color: {COLORS['background']};
                }}
            """)
            
        error_dialog.exec()

    def show_custom_success(self, title, message, detail):
        success_dialog = QDialog(self)
        success_dialog.setWindowTitle(title)
        success_dialog.setFixedWidth(400)
        
        layout = QVBoxLayout(success_dialog)
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)

        # Icono de éxito
        icon_label = QLabel()
        icon_label.setText("✅")
        icon_label.setStyleSheet("""
            QLabel {
                color: #4CAF50;
                font-size: 48px;
            }
        """)
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(icon_label)

        # Mensaje principal
        message_label = QLabel(message)
        message_label.setWordWrap(True)
        message_label.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text']};
                font-size: 16px;
                font-weight: bold;
            }}
        """)
        message_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(message_label)

        # Detalle
        detail_label = QLabel(detail)
        detail_label.setWordWrap(True)
        detail_label.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text_secondary']};
                font-size: 14px;
            }}
        """)
        detail_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(detail_label)

        # Botón de cerrar
        close_btn = QPushButton("Continuar")
        close_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['success']};
                color: {COLORS['text']};
                border: none;
                padding: 12px;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
                min-width: 100px;
            }}
            QPushButton:hover {{
                background-color: #43A047;
            }}
            QPushButton:pressed {{
                background-color: #388E3C;
            }}
        """)
        close_btn.clicked.connect(success_dialog.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignmentFlag.AlignCenter)

        # Estilo general del diálogo
        success_dialog.setStyleSheet(f"""
            QDialog {{
                background-color: {COLORS['background']};
            }}
        """)

        success_dialog.exec()

    def show_custom_warning(self, title, message, detail):
        warning_dialog = QDialog(self)
        warning_dialog.setWindowTitle(title)
        warning_dialog.setFixedWidth(400)
        
        layout = QVBoxLayout(warning_dialog)
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)

        # Icono de advertencia
        icon_label = QLabel()
        icon_label.setText("⚠️")
        icon_label.setStyleSheet("""
            QLabel {
                color: #FFA500;
                font-size: 48px;
            }
        """)
        icon_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(icon_label)

        # Mensaje principal
        message_label = QLabel(message)
        message_label.setWordWrap(True)
        message_label.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text']};
                font-size: 16px;
                font-weight: bold;
            }}
        """)
        message_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(message_label)

        # Detalle
        detail_label = QLabel(detail)
        detail_label.setWordWrap(True)
        detail_label.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text_secondary']};
                font-size: 14px;
            }}
        """)
        detail_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(detail_label)

        # Botón de cerrar
        close_btn = QPushButton("Entendido")
        close_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['primary']};
                color: {COLORS['text']};
                border: none;
                padding: 12px;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
                min-width: 100px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['primary_dark']};
            }}
            QPushButton:pressed {{
                background-color: {COLORS['primary']};
            }}
        """)
        close_btn.clicked.connect(warning_dialog.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignmentFlag.AlignCenter)

        # Estilo general del diálogo
        warning_dialog.setStyleSheet(f"""
                QDialog {{
                    background-color: {COLORS['background']};
                }}
            """)
            
        warning_dialog.exec()

    def validate_username(self):
        text = self.username_input.text()
        validation_result = {
            'valid': True,
            'message': "",
            'details': []
        }

        # Validaciones con mensajes detallados
        if len(text) < 4:
            validation_result['valid'] = False
            validation_result['message'] = "Usuario demasiado corto"
            validation_result['details'] = [
                "Mínimo 4 caracteres requeridos",
                f"Actualmente: {len(text)} caracteres"
            ]
        elif not text.isalnum():
            validation_result['valid'] = False
            validation_result['message'] = "Caracteres no permitidos"
            validation_result['details'] = [
                "Solo se permiten letras y números",
                "No usar espacios ni caracteres especiales"
            ]
        else:
            validation_result['message'] = "Usuario válido"
            validation_result['details'] = [
                "Formato correcto",
                "Longitud adecuada"
            ]

        self.update_field_status(
            self.username_status, 
            self.username_input,
            validation_result
        )
        return validation_result['valid']

    def validate_password(self):
        text = self.password_input.text()
        validation_result = {
            'valid': True,
            'message': "",
            'details': []
        }

        # Lista de verificación de requisitos
        requirements = []
        requirements.append(('length', len(text) >= 8, "Mínimo 8 caracteres"))
        requirements.append(('uppercase', any(c.isupper() for c in text), "Una mayúscula"))
        requirements.append(('digit', any(c.isdigit() for c in text), "Un número"))

        # Verificar requisitos
        failed_requirements = [req[2] for req in requirements if not req[1]]

        if failed_requirements:
            validation_result['valid'] = False
            validation_result['message'] = "Contraseña débil"
            validation_result['details'] = [
                "Requisitos faltantes:",
                *failed_requirements
            ]
        else:
            validation_result['message'] = "Contraseña segura"
            validation_result['details'] = [
                "Cumple todos los requisitos",
                "✓ Longitud adecuada",
                "✓ Incluye mayúsculas",
                "✓ Incluye números"
            ]

        self.update_field_status(
            self.password_status,
            self.password_input,
            validation_result
        )
        return validation_result['valid']

    def update_field_status(self, status_label, input_field, validation):
        # Actualizar el icono y tooltip
        icon = "✓" if validation['valid'] else "⚠️"
        
        # Crear tooltip detallado
        tooltip = f"""
        <h3 style='color: {"#4CAF50" if validation["valid"] else "#EF5350"};'>
            {validation['message']}
        </h3>
        <ul style='margin: 5px 0;'>
            {"".join(f"<li>{detail}</li>" for detail in validation['details'])}
        </ul>
        """
        
        # Actualizar el estilo del campo según validación
        input_field.setStyleSheet(f"""
            QLineEdit {{
                padding: 10px 12px;
                border: 2px solid {COLORS['success'] if validation['valid'] else COLORS['error']};
                border-radius: 6px;
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                font-size: 14px;
                transition: all 0.3s;
            }}
            QLineEdit:focus {{
                border: 2px solid {COLORS['primary']};
            }}
            QLineEdit::placeholder {{
                color: {COLORS['text_secondary']};
                font-size: 15px;
                opacity: 0.95;
                font-weight: 450;
                letter-spacing: 0.4px;
            }}
        """)

        # Actualizar el label de estado
        status_label.setText(icon)
        status_label.setToolTip(tooltip)
        status_label.setStyleSheet("""
            QLabel {
                margin-left: 5px;
                font-size: 16px;
                padding: 5px;
            }
            QToolTip {
                background-color: #2b2b2b;
                color: white;
                border: 1px solid #555555;
                border-radius: 4px;
                padding: 8px;
                font-size: 12px;
            }
        """)

    def toggle_password_visibility(self):
        """Alternar visibilidad de la contraseña"""
        if self.password_input.echoMode() == QLineEdit.EchoMode.Password:
            self.password_input.setEchoMode(QLineEdit.EchoMode.Normal)
            self.toggle_password_btn.setText("🔓")  # Candado abierto
        else:
            self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
            self.toggle_password_btn.setText("🔒")  # Candado cerrado

    def toggle_confirm_visibility(self):
        """Alternar visibilidad de la confirmación de contraseña"""
        if self.confirm_password_input.echoMode() == QLineEdit.EchoMode.Password:
            self.confirm_password_input.setEchoMode(QLineEdit.EchoMode.Normal)
            self.toggle_confirm_btn.setText("🔓")  # Candado abierto
        else:
            self.confirm_password_input.setEchoMode(QLineEdit.EchoMode.Password)
            self.toggle_confirm_btn.setText("🔒")  # Candado cerrado

class AdminPanel(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.current_admin = parent.email  # Agregamos esta línea
        self.setWindowTitle("Panel de Administración")
        self.setMinimumWidth(1200)
        self.setMinimumHeight(700)
        
        # Inicializar atributos de clase
        self.search_input = None
        self.role_filter = None
        self.user_table = None
        
        # Inicializar UI
        self._init_ui()
        
        # Cargar datos
        try:
            self.load_users()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar usuarios: {str(e)}")

    def _init_ui(self):
        """Inicializa y configura la interfaz de usuario"""
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(30, 30, 30, 30)

        # Header con botón de solicitudes
        header_widget = self._create_header()
        
        # Botón de solicitudes con badge de notificación
        self.requests_btn = QPushButton("🔔 Solicitudes")
        self.requests_badge = QLabel("0")
        self.requests_badge.setStyleSheet("""
            QLabel {
                background-color: #FF5252;
                color: white;
                border-radius: 10px;
                padding: 2px 6px;
                font-size: 12px;
                font-weight: bold;
            }
        """)
        self.requests_badge.hide()
        
        # Layout para el botón y su badge
        request_layout = QHBoxLayout()
        request_layout.addWidget(self.requests_btn)
        request_layout.addWidget(self.requests_badge)
        request_layout.setAlignment(Qt.AlignmentFlag.AlignRight)
        
        self.requests_btn.clicked.connect(self.show_pending_requests)
        self.requests_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 8px 15px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        
        main_layout.addLayout(request_layout)

        # Búsqueda y filtros
        search_widget = self._create_search_section()
        main_layout.addWidget(search_widget)

        # Tabla con sombra
        table_container = self._create_table_container()
        main_layout.addWidget(table_container)

        # Botones de acción
        button_container = self._create_button_container()
        main_layout.addWidget(button_container)

        self.setStyleSheet(f"""
            QDialog {{
                background-color: {COLORS['background']};
            }}
        """)

    def _create_header(self):
        """Crea la sección del encabezado con estadísticas"""
        header = QWidget()
        layout = QHBoxLayout(header)
        layout.setContentsMargins(0, 0, 0, 20)
        
        # Título con icono
        title = QLabel("👥 Administración de Usuarios")
        title.setStyleSheet(f"""
            QLabel {{
                color: {COLORS['text']};
                font-size: 24px;
                font-weight: bold;
            }}
        """)
        
        layout.addWidget(title)
        layout.addStretch()
        
        return header

    def _create_search_section(self):
        """Crea la sección de búsqueda y filtros"""
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        
        # Barra de búsqueda
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 Buscar usuarios...")
        self.search_input.textChanged.connect(self._filter_users)
        self.search_input.setStyleSheet(f"""
            QLineEdit {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                border: none;
                border-radius: 5px;
                padding: 8px;
                font-size: 14px;
            }}
        """)
        
        # Filtro por departamento
        self.dept_filter = QComboBox()
        self.dept_filter.addItem("Todos los departamentos")
        # Obtener departamentos de la base de datos
        try:
            departamentos = DatabaseManager.execute_query("""
                SELECT DISTINCT departamento 
                FROM usuario 
                WHERE departamento IS NOT NULL 
                ORDER BY departamento
            """)
            for dept in departamentos:
                if dept['departamento']:  # Asegurarse de que no sea None
                    self.dept_filter.addItem(dept['departamento'])
        except Exception as e:
            print(f"Error al cargar departamentos: {str(e)}")

        self.dept_filter.currentTextChanged.connect(self._filter_users)
        self.dept_filter.setStyleSheet(f"""
            QComboBox {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                border: none;
                border-radius: 5px;
                padding: 8px;
                min-width: 200px;
            }}
            QComboBox:hover {{
                border: 1px solid {COLORS['primary']};
            }}
            QComboBox::drop-down {{
                border: none;
            }}
            QComboBox::down-arrow {{
                image: url(down_arrow.png);
                width: 12px;
                height: 12px;
            }}
        """)
        
        layout.addWidget(self.search_input, stretch=2)
        layout.addWidget(self.dept_filter, stretch=1)
        return container

    def _create_table_container(self):
        """Crea el contenedor de la tabla con efectos visuales"""
        container = QWidget()
        layout = QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        
        self.user_table = QTableWidget()
        self.user_table.setColumnCount(5)
        self.user_table.setHorizontalHeaderLabels([
            "Usuario", "Email", "Departamento", "Rol", "Acciones"
        ])
        
        # Establecer altura de las filas
        self.user_table.verticalHeader().setDefaultSectionSize(33)  # Aumenta la altura de las filas a 60px
        
        # Opcional: Establecer altura mínima de las filas
        self.user_table.verticalHeader().setMinimumSectionSize(33)
        
        # Configurar el estiramiento de las columnas
        header = self.user_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)  # Usuario
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Email
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)  # Departamento
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)  # Rol
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)  # Acciones
        
        # Establecer anchos específicos para las columnas
        self.user_table.setColumnWidth(0, 150)  # Usuario
        self.user_table.setColumnWidth(2, 200)  # Departamento
        self.user_table.setColumnWidth(3, 100)  # Rol
        self.user_table.setColumnWidth(4, 100)  # Acciones
        
        # Configurar estilos y comportamiento de la tabla
        self.user_table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                border: none;
                border-radius: 8px;
                gridline-color: {COLORS['primary']};
            }}
            QTableWidget::item {{
                padding: 8px;
                border-bottom: 1px solid {COLORS['primary']};
                height: 60px;  /* Altura adicional para los items */
            }}
            QTableWidget::item:selected {{
                background-color: {COLORS['primary']};
                color: {COLORS['text']};
            }}
            QHeaderView::section {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                padding: 12px 8px;  /* Aumentado el padding vertical del encabezado */
                border: none;
                border-bottom: 2px solid {COLORS['primary']};
                font-weight: bold;
            }}
        """)
        
        # Configuraciones adicionales de la tabla
        self.user_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.user_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.user_table.verticalHeader().setVisible(False)
        self.user_table.setShowGrid(True)
        self.user_table.setAlternatingRowColors(True)
        
        layout.addWidget(self.user_table)
        return container

    def _create_button_container(self):
        """Crea el contenedor con los botones de acción"""
        container = QWidget()
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 20, 0, 0)
        
        refresh_button = QPushButton("🔄 Actualizar")
        refresh_button.clicked.connect(self.load_users)
        refresh_button.setStyleSheet(self._get_button_style())
        
        add_button = QPushButton("➕ Nuevo Usuario")
        add_button.clicked.connect(self.add_user)
        add_button.setStyleSheet(self._get_button_style('primary'))
        
        save_button = QPushButton("💾 Guardar Cambios")
        save_button.clicked.connect(self.save_changes)
        save_button.setStyleSheet(self._get_button_style('success'))
        
        layout.addWidget(add_button)
        layout.addStretch()
        layout.addWidget(refresh_button)
        layout.addWidget(save_button)
        
        return container

    def _configure_table_dimensions(self):
        """Configura las dimensiones de las columnas de la tabla"""
        header = self.user_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)  # Usuario
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Email
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Fixed)  # Departamento
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Fixed)  # Rol Actual
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Fixed)  # Nuevo Rol
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)  # Acciones
        
        self.user_table.setColumnWidth(0, 100)
        self.user_table.setColumnWidth(1, 100)
        self.user_table.setColumnWidth(2, 100)
        self.user_table.setColumnWidth(3, 100)
        self.user_table.setColumnWidth(4, 100)
        self.user_table.setColumnWidth(5, 100)

    def load_users(self):
        """Carga los usuarios desde la base de datos"""
        try:
            users = DatabaseManager.execute_query("""
                SELECT nombreusuario, email, departamento, rol 
                FROM usuario 
                ORDER BY departamento, nombreusuario
            """)
            
            self.user_table.setRowCount(len(users))
            
            for i, user in enumerate(users):
                self._populate_user_row(i, user)
                
            # Actualizar la lista de departamentos en el filtro
            self._update_department_filter()
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar usuarios: {str(e)}")

    def _populate_user_row(self, row, user):
        """Rellena una fila de la tabla con los datos del usuario"""
        # Usuario
        self.user_table.setItem(row, 0, QTableWidgetItem(user['nombreusuario']))
        
        # Email
        self.user_table.setItem(row, 1, QTableWidgetItem(user['email']))
        
        # Departamento
        self.user_table.setItem(row, 2, QTableWidgetItem(user['departamento']))
        
        # Rol
        self.user_table.setItem(row, 3, QTableWidgetItem(user['rol']))
        
        # Botones de acción
        action_widget = QWidget()
        action_layout = QHBoxLayout(action_widget)
        action_layout.setContentsMargins(0, 0, 0, 0)  # Eliminados los márgenes
        action_layout.setSpacing(2)  # Reducido el espacio entre botones
        
        # Botón editar
        edit_button = QPushButton("✏️")
        edit_button.setToolTip("Editar usuario")
        edit_button.clicked.connect(lambda: self.edit_user(user))
        edit_button.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
                padding: 1px;
                font-size: 16px;
                min-width: 25px;
                min-height: 25px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2196F3;
                border-radius: 4px;
                color: white;
            }
        """)
        
        # Botón eliminar
        delete_button = QPushButton("🗑️")
        delete_button.setToolTip("Eliminar usuario")
        delete_button.clicked.connect(lambda: self.delete_user(user))
        delete_button.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
                padding: 1px;
                font-size: 16px;
                min-width: 25px;
                min-height: 25px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #EF5350;
                border-radius: 4px;
                color: white;
            }
        """)
        
        action_layout.addWidget(edit_button)
        action_layout.addWidget(delete_button)
        action_layout.addStretch()
        
        self.user_table.setCellWidget(row, 4, action_widget)

        # Ajustar el ancho de la columna de acciones
        self.user_table.setColumnWidth(4, 70)  # Reducido el ancho de la columna

    def _create_action_buttons(self, user):
        """Crea los botones de acción para un usuario"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(5, 0, 5, 0)
        
        edit_btn = QPushButton("✏️")
        edit_btn.setToolTip("Editar usuario")
        edit_btn.clicked.connect(lambda: self.edit_user(user))
        edit_btn.setStyleSheet(self._get_action_button_style())
        
        delete_btn = QPushButton("🗑️")
        delete_btn.setToolTip("Eliminar usuario")
        delete_btn.clicked.connect(lambda: self.delete_user(user))
        delete_btn.setStyleSheet(self._get_action_button_style('danger'))
        
        layout.addWidget(edit_btn)
        layout.addWidget(delete_btn)
        
        return widget

    def _filter_users(self):
        """Filtra los usuarios según la búsqueda y el departamento seleccionado"""
        search_text = self.search_input.text().lower()
        selected_dept = self.dept_filter.currentText()
        
        for row in range(self.user_table.rowCount()):
            show_row = True
            username = self.user_table.item(row, 0).text().lower()
            email = self.user_table.item(row, 1).text().lower()
            dept = self.user_table.item(row, 2).text()
            
            # Filtrar por texto de búsqueda (en usuario y email)
            if search_text and search_text not in username and search_text not in email:
                show_row = False
                
            # Filtrar por departamento
            if selected_dept != "Todos los departamentos" and dept != selected_dept:
                show_row = False
                
            self.user_table.setRowHidden(row, not show_row)

    def add_user(self):
        """Abre el diálogo para agregar un nuevo usuario"""
        # Implementar la lógica para agregar usuario
        pass

    def edit_user(self, user):
        """Abre el diálogo para editar un usuario"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Editar Usuario")
        dialog.setMinimumWidth(400)
        layout = QFormLayout(dialog)

        # Campos existentes
        email_input = QLineEdit(user['email'])
        email_input.setStyleSheet(f"""
            QLineEdit {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                border: 1px solid {COLORS['primary']};
                border-radius: 5px;
                padding: 8px;
            }}
        """)

        # Campos para cambiar contraseña
        password_input = QLineEdit()
        password_input.setEchoMode(QLineEdit.EchoMode.Password)
        password_input.setPlaceholderText("Nueva contraseña (dejar vacío para no cambiar)")
        password_input.setStyleSheet(email_input.styleSheet())

        confirm_password = QLineEdit()
        confirm_password.setEchoMode(QLineEdit.EchoMode.Password)
        confirm_password.setPlaceholderText("Confirmar nueva contraseña")
        confirm_password.setStyleSheet(email_input.styleSheet())

        # Resto de campos existentes
        dept_combo = QComboBox()
        dept_combo.setStyleSheet(self._get_combo_style())
        
        try:
            departamentos = DatabaseManager.execute_query("""
                SELECT nombre_departamento 
                FROM departamento 
                ORDER BY nombre_departamento
            """)
            
            for dept in departamentos:
                dept_combo.addItem(dept['nombre_departamento'])
            
            current_dept_index = dept_combo.findText(user['departamento'])
            if current_dept_index >= 0:
                dept_combo.setCurrentIndex(current_dept_index)
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar departamentos: {str(e)}")

        role_combo = QComboBox()
        role_combo.addItems(["usuario", "recepcionista", "admin"])
        role_combo.setCurrentText(user['rol'])
        role_combo.setStyleSheet(self._get_combo_style())

        # Agregar campos al layout
        layout.addRow("Email:", email_input)
        layout.addRow("Departamento:", dept_combo)
        layout.addRow("Rol:", role_combo)
        layout.addRow("Nueva Contraseña:", password_input)
        layout.addRow("Confirmar Contraseña:", confirm_password)

        button_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | 
            QDialogButtonBox.StandardButton.Cancel
        )
        
        button_box.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['primary']};
                color: {COLORS['text']};
                border: none;
                border-radius: 5px;
                padding: 8px 16px;
                min-width: 80px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['primary_dark']};
            }}
        """)

        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addRow(button_box)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            try:
                # Verificar si se quiere cambiar la contraseña
                if password_input.text():
                    if password_input.text() != confirm_password.text():
                        QMessageBox.critical(
                            self,
                            "Error",
                            "Las contraseñas no coinciden"
                        )
                        return
                    
                    # Generar nuevo salt y hash para la nueva contraseña
                    new_salt = DatabaseManager.generate_salt()
                    new_password_hash = DatabaseManager.hash_password(password_input.text(), new_salt)
                    
                    # Actualizar usuario con nueva contraseña
                    DatabaseManager.execute_query("""
                        UPDATE usuario 
                        SET email = %s, 
                            departamento = %s, 
                            rol = %s,
                            password_hash = %s,
                            salt = %s
                        WHERE nombreusuario = %s
                    """, (
                        email_input.text(),
                        dept_combo.currentText(),
                        role_combo.currentText(),
                        new_password_hash,
                        new_salt,
                        user['nombreusuario']
                    ))
                else:
                    # Actualizar usuario sin cambiar contraseña
                    DatabaseManager.execute_query("""
                        UPDATE usuario 
                        SET email = %s, 
                            departamento = %s, 
                            rol = %s 
                        WHERE nombreusuario = %s
                    """, (
                        email_input.text(),
                        dept_combo.currentText(),
                        role_combo.currentText(),
                        user['nombreusuario']
                    ))
                
                self.load_users()
                QMessageBox.information(
                    self,
                    "Éxito",
                    "Usuario actualizado correctamente"
                )
                
            except Exception as e:
                QMessageBox.critical(
                    self,
                    "Error",
                    f"Error al actualizar usuario: {str(e)}"
                )

    def delete_user(self, user):
        """Elimina un usuario después de confirmar"""
        reply = QMessageBox.question(
            self,
            "Confirmar Eliminación",
            f"¿Está seguro de que desea eliminar al usuario {user['nombreusuario']}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                DatabaseManager.execute_query(
                    "DELETE FROM usuario WHERE nombreusuario = %s",
                    (user['nombreusuario'],))
                self.load_users()
                QMessageBox.information(self, "Éxito", "Usuario eliminado correctamente")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al eliminar usuario: {str(e)}")

    def save_changes(self):
        """Guarda los cambios en los roles de usuarios"""
        try:
            changes = []
            for row in range(self.user_table.rowCount()):
                username = self.user_table.item(row, 0).text()
                current_role = self.user_table.item(row, 3).text()
                new_role = self.user_table.cellWidget(row, 4).currentText()
                
                if current_role != new_role:
                    changes.append((username, new_role))
            
            if not changes:
                QMessageBox.information(self, "Info", "No hay cambios para guardar")
                return
            
            for username, new_role in changes:
                DatabaseManager.execute_query(
                    "UPDATE usuario SET rol = %s WHERE nombreusuario = %s",
                    (new_role, username)
                )
            
            self.load_users()
            QMessageBox.information(self, "Éxito", "Cambios guardados correctamente")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al guardar cambios: {str(e)}")

    # Métodos de estilo
    @staticmethod
    def _get_button_style(button_type='default'):
        styles = {
            'default': f"""
                QPushButton {{
                    background-color: {COLORS['surface']};
                    color: {COLORS['text']};
                    border: 2px solid {COLORS['primary']};
                    padding: 8px 15px;
                    border-radius: 5px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: {COLORS['primary']};
                }}
            """,
            'primary': f"""
                QPushButton {{
                    background-color: {COLORS['primary']};
                    color: {COLORS['text']};
                    border: none;
                    padding: 8px 15px;
                    border-radius: 5px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: {COLORS['primary_dark']};
                }}
            """,
            'success': f"""
                QPushButton {{
                    background-color: {COLORS['success']};
                    color: {COLORS['text']};
                    border: none;
                    padding: 8px 15px;
                    border-radius: 5px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: #388E3C;
                }}
            """
        }
        return styles.get(button_type, styles['default'])

    @staticmethod
    def _get_action_button_style(button_type='default'):
        base_style = f"""
            QPushButton {{
                background-color: transparent;
                border: none;
                border-radius: 3px;
                padding: 5px;
                font-size: 16px;
            }}
        """
        
        if button_type == 'danger':
            base_style += f"""
                QPushButton:hover {{
                    background-color: {COLORS['error']};
                }}
            """
        else:
            base_style += f"""
                QPushButton:hover {{
                    background-color: {COLORS['primary']};
                }}
            """
        return base_style

    @staticmethod
    def _get_combo_style():
        return f"""
            QComboBox {{
                background-color: {COLORS['surface']};
                color: {COLORS['text']};
                border: 1px solid {COLORS['primary']};
                border-radius: 3px;
                padding: 5px;
            }}
            QComboBox:hover {{
                border-color: {COLORS['primary_light']};
            }}
        """

    def _update_department_filter(self):
        """Actualiza la lista de departamentos en el filtro"""
        current_dept = self.dept_filter.currentText()
        self.dept_filter.clear()
        self.dept_filter.addItem("Todos los departamentos")
        
        try:
            departamentos = DatabaseManager.execute_query("""
                SELECT DISTINCT departamento 
                FROM usuario 
                WHERE departamento IS NOT NULL 
                ORDER BY departamento
            """)
            
            for dept in departamentos:
                if dept['departamento']:
                    self.dept_filter.addItem(dept['departamento'])
                    
            # Restaurar la selección anterior si aún existe
            index = self.dept_filter.findText(current_dept)
            if index >= 0:
                self.dept_filter.setCurrentIndex(index)
                
        except Exception as e:
            print(f"Error al actualizar departamentos: {str(e)}")

    def _setup_theme_selector(self):
        themes = {
            "Dark": {
                "background": "#1e1e1e",
                "surface": "#2b2b2b",
                "primary": "#1976D2",
                "text": "#ffffff"
            },
            "Light": {
                "background": "#ffffff",
                "surface": "#f5f5f5",
                "primary": "#2196F3",
                "text": "#000000"
            },
            "Night": {
                "background": "#000000",
                "surface": "#121212",
                "primary": "#BB86FC",
                "text": "#ffffff"
            }
        }
        
        theme_menu = QMenu("🎨 Temas")
        for theme_name in themes:
            action = theme_menu.addAction(theme_name)
            action.triggered.connect(lambda checked, t=theme_name: self._apply_theme(themes[t]))

    def _create_notification_center(self):
        notification_widget = QWidget()
        layout = QVBoxLayout(notification_widget)
        
        notifications = [
            ("⚠️", "5 usuarios pendientes de aprobación"),
            ("🔒", "3 intentos fallidos de inicio de sesión"),
            ("📊", "Reporte mensual disponible"),
            ("🔄", "Actualización del sistema pendiente")
        ]
        
        for icon, text in notifications:
            note = QPushButton(f"{icon} {text}")
            note.setStyleSheet("""
                QPushButton {
                    background-color: #2b2b2b;
                    border: none;
                    border-radius: 5px;
                    padding: 10px;
                    text-align: left;
                    margin: 2px;
                }
                QPushButton:hover {
                    background-color: #333333;
                }
            """)
            layout.addWidget(note)

    def show_pending_requests(self):
        dialog = PendingRequestsDialog(self)
        dialog.exec()

    def update_pending_count(self):
        try:
            count = DatabaseManager.execute_query("""
                SELECT COUNT(*) as count 
                FROM usuario 
                WHERE estado = 'pendiente'
            """)[0]['count']
            
            if count > 0:
                self.requests_badge.setText(str(count))
                self.requests_badge.show()
            else:
                self.requests_badge.hide()
                
        except Exception as e:
            print(f"Error al actualizar contador de solicitudes: {str(e)}")

class PendingRequestsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Solicitudes Pendientes")
        self.setMinimumWidth(800)
        self.setMinimumHeight(500)
        
        layout = QVBoxLayout(self)
        
        # Tabla de solicitudes
        self.requests_table = QTableWidget()
        self.requests_table.setColumnCount(7)
        self.requests_table.setHorizontalHeaderLabels([
            "Usuario", "Email", "Departamento", 
            "Fecha Solicitud", "Estado", "Acciones", "Detalles"
        ])
        
        # Configurar la tabla
        self.requests_table.setStyleSheet("""
            QTableWidget {
                background-color: #2b2b2b;
                border: none;
                border-radius: 8px;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #333333;
            }
        """)
        
        layout.addWidget(self.requests_table)
        
        self.load_pending_requests()

    def load_pending_requests(self):
        try:
            requests = DatabaseManager.execute_query("""
                SELECT nombreusuario, email, departamento, fecha_solicitud, estado
                FROM usuario 
                WHERE estado = 'pendiente'
                ORDER BY fecha_solicitud DESC
            """)
            
            self.requests_table.setRowCount(len(requests))
            
            for i, req in enumerate(requests):
                self.requests_table.setItem(i, 0, QTableWidgetItem(req['nombreusuario']))
                self.requests_table.setItem(i, 1, QTableWidgetItem(req['email']))
                self.requests_table.setItem(i, 2, QTableWidgetItem(req['departamento']))
                
                # Agregar verificación para fecha_solicitud
                fecha = req['fecha_solicitud']
                fecha_str = fecha.strftime("%Y-%m-%d %H:%M") if fecha else "Sin fecha"
                self.requests_table.setItem(i, 3, QTableWidgetItem(fecha_str))
                
                # Estado con color
                estado_item = QTableWidgetItem("PENDIENTE")
                estado_item.setForeground(QColor("#FFC107"))
                self.requests_table.setItem(i, 4, estado_item)
                
                # Botones de acción
                action_widget = QWidget()
                action_layout = QHBoxLayout(action_widget)
                action_layout.setContentsMargins(0, 0, 0, 0)
                
                approve_btn = QPushButton("✅")
                reject_btn = QPushButton("❌")
                
                approve_btn.clicked.connect(lambda _, u=req['nombreusuario']: self.approve_user(u))
                reject_btn.clicked.connect(lambda _, u=req['nombreusuario']: self.reject_user(u))
                
                for btn in [approve_btn, reject_btn]:
                    btn.setStyleSheet("""
                        QPushButton {
                            background-color: transparent;
                            border: none;
                            font-size: 16px;
                            padding: 5px;
                        }
                        QPushButton:hover {
                            background-color: #333333;
                            border-radius: 3px;
                        }
                    """)
                
                action_layout.addWidget(approve_btn)
                action_layout.addWidget(reject_btn)
                self.requests_table.setCellWidget(i, 5, action_widget)
                
                # Botón de detalles
                details_btn = QPushButton("👁️")
                details_btn.clicked.connect(lambda _, u=req: self.show_user_details(u))
                details_btn.setStyleSheet("""
                    QPushButton {
                        background-color: transparent;
                        border: none;
                        font-size: 16px;
                        padding: 5px;
                    }
                    QPushButton:hover {
                        background-color: #333333;
                        border-radius: 3px;
                    }
                """)
                self.requests_table.setCellWidget(i, 6, details_btn)
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar solicitudes: {str(e)}")

    def approve_user(self, username):
        try:
            # Obtener información del usuario
            user_info = DatabaseManager.execute_query("""
                SELECT email FROM usuario WHERE nombreusuario = %s
            """, (username,))[0]
            
            # Actualizar estado en la base de datos
            DatabaseManager.execute_query("""
                UPDATE usuario 
                SET estado = 'aprobado',
                    fecha_aprobacion = CURRENT_TIMESTAMP,
                    aprobado_por = %s
                WHERE nombreusuario = %s
            """, (self.parent().current_admin, username))
            
            # Enviar notificación por email
            if EmailNotifier.send_approval_notification(user_info['email'], username):
                QMessageBox.information(
                    self,
                    "Éxito",
                    "Usuario aprobado y notificación enviada correctamente"
                )
            else:
                QMessageBox.warning(
                    self,
                    "Advertencia",
                    "Usuario aprobado pero hubo un problema al enviar la notificación"
                )
            
            self.load_pending_requests()
            self.parent().update_pending_count()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al aprobar usuario: {str(e)}")

    def reject_user(self, username):
        reason, ok = QInputDialog.getText(
            self, 
            "Motivo de Rechazo",
            "Por favor, indique el motivo del rechazo:"
        )
        
        if ok and reason:
            try:
                # Obtener información del usuario
                user_info = DatabaseManager.execute_query("""
                    SELECT email FROM usuario WHERE nombreusuario = %s
                """, (username,))[0]
                
                # Actualizar estado en la base de datos
                DatabaseManager.execute_query("""
                    UPDATE usuario 
                    SET estado = 'rechazado',
                        fecha_rechazo = CURRENT_TIMESTAMP,
                        rechazado_por = %s,
                        motivo_rechazo = %s
                    WHERE nombreusuario = %s
                """, (self.parent().current_admin, reason, username))
                
                # Enviar notificación por email
                if EmailNotifier.send_rejection_notification(user_info['email'], username, reason):
                    QMessageBox.information(
                        self,
                        "Éxito",
                        "Usuario rechazado y notificación enviada correctamente"
                    )
                else:
                    QMessageBox.warning(
                        self,
                        "Advertencia",
                        "Usuario rechazado pero hubo un problema al enviar la notificación"
                    )
                
                self.load_pending_requests()
                self.parent().update_pending_count()
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al rechazar usuario: {str(e)}")

    def show_user_details(self, user):
        # Implementar la lógica para mostrar detalles del usuario
        pass

    def send_approval_notification(self, username):
        # Implementar la lógica para enviar notificación de aprobación
        pass

    def send_rejection_notification(self, username, reason):
        # Implementar la lógica para enviar notificación de rechazo
        pass

# Función auxiliar para crear el estilo de input (para reutilizar en ambas clases)
def create_input_style():
    return """
        QLineEdit {
            background-color: black;
            color: white;
            border: 1px solid white;
        }
    """

def create_button_style():
    return """
        QPushButton {
            background-color: black;
            color: white;
            border: 1px solid white;
        }
    """

def create_label_style():
    return """
        QLabel {
            color: white;
        }
    """

class PDFManager:
    CHUNK_SIZE = 8192  # 8KB chunks para lectura/escritura
    
    @staticmethod
    def save_pdf(id_documento, pdf_data):
        """Guarda un PDF en la base de datos de forma optimizada"""
        try:
            query = """
                UPDATE documento 
                SET archivo_pdf = %s 
                WHERE id_documento = %s
            """
            DatabaseManager.execute_query(query, (pdf_data, id_documento))
            
        except Exception as e:
            raise Exception(f"Error guardando PDF: {str(e)}")

    @staticmethod
    def get_pdf(id_documento):
        """Obtiene un PDF de la base de datos de forma optimizada"""
        try:
            query = """
                SELECT archivo_pdf 
                FROM documento 
                WHERE id_documento = %s 
                AND archivo_pdf IS NOT NULL
            """
            result = DatabaseManager.execute_query(query, (id_documento,))
            
            if not result:
                raise ValueError("PDF no encontrado")
                
            return result[0]['archivo_pdf']
            
        except Exception as e:
            raise Exception(f"Error obteniendo PDF: {str(e)}")

    @staticmethod
    def save_pdf_to_file(pdf_data, file_path):
        """Guarda un PDF en el sistema de archivos de forma optimizada"""
        try:
            with open(file_path, 'wb') as f:
                for i in range(0, len(pdf_data), PDFManager.CHUNK_SIZE):
                    chunk = pdf_data[i:i + PDFManager.CHUNK_SIZE]
                    f.write(chunk)
                    
        except Exception as e:
            raise Exception(f"Error guardando PDF en archivo: {str(e)}")

class UIManager:
    # Cache para estilos y fuentes
    _style_cache = {}
    _font_cache = {}
    
    @classmethod
    def get_style(cls, key, params=None):
        """Obtiene estilos cacheados para mejor rendimiento"""
        cache_key = f"{key}_{str(params)}"
        if cache_key not in cls._style_cache:
            cls._style_cache[cache_key] = cls._generate_style(key, params)
        return cls._style_cache[cache_key]

    @classmethod
    def get_font(cls, size, bold=False):
        """Obtiene fuentes cacheadas"""
        cache_key = f"{size}_{bold}"
        if cache_key not in cls._font_cache:
            font = QFont("Segoe UI", size)
            font.setBold(bold)
            cls._font_cache[cache_key] = font
        return cls._font_cache[cache_key]

    @staticmethod
    def create_button(text, icon_path=None, tooltip=None):
        """Crea botones con estilo consistente"""
        button = QPushButton(text)
        if icon_path:
            button.setIcon(QIcon(resource_path(icon_path)))
        if tooltip:
            button.setToolTip(tooltip)
        button.setStyleSheet(UIManager.get_style('button'))
        button.setCursor(Qt.CursorShape.PointingHandCursor)
        return button

    @staticmethod
    def create_tree_widget(headers):
        """Crea un TreeWidget optimizado"""
        tree = QTreeWidget()
        tree.setHeaderLabels(headers)
        tree.setAlternatingRowColors(True)
        tree.setStyleSheet(UIManager.get_style('tree'))
        tree.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        tree.setSortingEnabled(True)
        return tree

    @staticmethod
    def _generate_style(key, params=None):
        """Genera estilos según el tipo de widget"""
        styles = {
            'button': f"""
                QPushButton {{
                    background-color: {COLORS['primary']};
                    color: {COLORS['text']};
                    border: none;
                    padding: 10px;
                    border-radius: 6px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: {COLORS['primary_dark']};
                }}
                QPushButton:pressed {{
                    background-color: {COLORS['primary']};
                }}
            """,
            'tree': f"""
                QTreeWidget {{
                    background-color: {COLORS['surface']};
                    border: none;
                    border-radius: 8px;
                }}
                QTreeWidget::item {{
                    padding: 5px;
                    border-bottom: 1px solid {COLORS['primary']};
                }}
                QTreeWidget::item:selected {{
                    background-color: {COLORS['primary']};
                    color: {COLORS['text']};
                }}
            """
        }
        return styles.get(key, "")

class SignalManager(QObject):
    # Señales personalizadas
    data_updated = pyqtSignal()
    error_occurred = pyqtSignal(str)
    progress_updated = pyqtSignal(int)
    
    _instance = None
    
    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    @classmethod
    def emit_data_updated(cls):
        """Emite señal de actualización de datos"""
        if cls._instance:
            cls._instance.data_updated.emit()

    @classmethod
    def emit_error(cls, message):
        """Emite señal de error"""
        if cls._instance:
            cls._instance.error_occurred.emit(message)

    @classmethod
    def emit_progress(cls, value):
        """Emite señal de progreso"""
        if cls._instance:
            cls._instance.progress_updated.emit(value)

class EmailManager:
    def __init__(self):
        # Configuración del servidor de correo (ejemplo con Gmail)
        self.smtp_server = "smtp.gmail.com"
        self.smtp_port = 587
        self.sender_email = "tu_correo@gmail.com"  # Correo desde donde se enviarán las notificaciones
        self.sender_password = "tu_contraseña_de_aplicacion"  # Contraseña de aplicación de Gmail
    
    def send_email(self, to_email, subject, html_content):
        try:
            # Crear mensaje
            message = MIMEMultipart('alternative')
            message['Subject'] = subject
            message['From'] = self.sender_email
            message['To'] = to_email

            # Convertir el contenido HTML
            html_part = MIMEText(html_content, 'html')
            message.attach(html_part)

            # Conectar y enviar
            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                server.starttls()
                server.login(self.sender_email, self.sender_password)
                server.send_message(message)
                
            return True
        except Exception as e:
            print(f"Error enviando email: {str(e)}")
            return False

class EmailNotifier:
    # Configuración para Gmail
    SMTP_SERVER = "smtp.gmail.com"
    SMTP_PORT = 587
    SENDER_EMAIL = "titoanthem1@gmail.com"
    APP_PASSWORD = "yfqo gwip fqke fbmc"

    @classmethod
    def send_registration_notification(cls, email):
        try:
            subject = "Solicitud de Registro - Corporación Isla de Maipo"
            
            # Crear el contenido HTML del correo
            html_content = f"""
            <div style="font-family: Arial, sans-serif; padding: 20px; max-width: 600px; margin: 0 auto;">
                <h2 style="color: #2196F3;">Solicitud de Registro Recibida</h2>
                
                <p>Estimado usuario,</p>
                
                <p>Hemos recibido su solicitud de registro en el Sistema de Gestión de la Corporación Isla de Maipo.</p>
                
                <div style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 20px 0;">
                    <p style="margin: 0;"><strong>Email registrado:</strong> {email}</p>
                    <p style="margin: 10px 0 0;"><strong>Estado:</strong> Pendiente de aprobación</p>
                </div>
                
                <p>Un administrador revisará su solicitud y recibirá una notificación cuando sea procesada.</p>
                
                <p style="color: #666; font-size: 12px; margin-top: 30px;">
                    Este es un correo automático, por favor no responda a este mensaje.<br>
                    Corporación Isla de Maipo © 2024
                </p>
            </div>
            """
            
            # Enviar el correo
            return cls._send_email(email, subject, html_content)
            
        except Exception as e:
            print(f"Error al enviar notificación de registro: {str(e)}")
            return False

    @classmethod
    def send_approval_notification(cls, email, username):
        try:
            subject = "Solicitud Aprobada - Corporación Isla de Maipo"
            
            html_content = f"""
            <div style="font-family: Arial, sans-serif; padding: 20px; max-width: 600px; margin: 0 auto;">
                <h2 style="color: #4CAF50;">¡Su solicitud ha sido aprobada!</h2>
                
                <p>Estimado {username},</p>
                
                <p>Nos complace informarle que su solicitud de registro ha sido aprobada.</p>
                
                <div style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 20px 0;">
                    <p style="margin: 0;"><strong>Usuario:</strong> {username}</p>
                    <p style="margin: 10px 0 0;"><strong>Email:</strong> {email}</p>
                </div>
                
                <p>Ya puede acceder al sistema utilizando sus credenciales.</p>
                
                <p style="color: #666; font-size: 12px; margin-top: 30px;">
                    Este es un correo automático, por favor no responda a este mensaje.<br>
                    Corporación Isla de Maipo © 2024
                </p>
            </div>
            """
            
            return cls._send_email(email, subject, html_content)
            
        except Exception as e:
            print(f"Error al enviar notificación de aprobación: {str(e)}")
            return False

    @classmethod
    def send_rejection_notification(cls, email, username, reason):
        try:
            subject = "Actualización de Solicitud - Corporación Isla de Maipo"
            
            html_content = f"""
            <div style="font-family: Arial, sans-serif; padding: 20px; max-width: 600px; margin: 0 auto;">
                <h2 style="color: #F44336;">Actualización de su Solicitud</h2>
                
                <p>Estimado {username},</p>
                
                <p>Lamentamos informarle que su solicitud de registro no ha sido aprobada.</p>
                
                <div style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 20px 0;">
                    <p style="margin: 0;"><strong>Motivo:</strong> {reason}</p>
                </div>
                
                <p>Si considera que esto es un error o necesita más información, por favor contacte al administrador.</p>
                
                <p style="color: #666; font-size: 12px; margin-top: 30px;">
                    Este es un correo automático, por favor no responda a este mensaje.<br>
                    Corporación Isla de Maipo © 2024
                </p>
            </div>
            """
            
            return cls._send_email(email, subject, html_content)
            
        except Exception as e:
            print(f"Error al enviar notificación de rechazo: {str(e)}")
            return False

    @classmethod
    def _send_email(cls, to_email, subject, html_content):
        try:
            # Crear mensaje
            message = MIMEMultipart('alternative')
            message['Subject'] = subject
            message['From'] = cls.SENDER_EMAIL
            message['To'] = to_email

            # Agregar contenido HTML
            html_part = MIMEText(html_content, 'html')
            message.attach(html_part)

            # Conectar y enviar
            with smtplib.SMTP(cls.SMTP_SERVER, cls.SMTP_PORT) as server:
                server.starttls()
                server.login(cls.SENDER_EMAIL, cls.APP_PASSWORD)
                server.send_message(message)
                
            print(f"Correo enviado exitosamente a {to_email}")
            return True
            
        except Exception as e:
            print(f"Error enviando email: {str(e)}")
            return False

def main():
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon(resource_path("isla_de_maipo.png")))
    
    # Mostramos el diálogo de login
    login = LoginDialog()
    if login.exec() == QDialog.DialogCode.Accepted:
        user_role = login.get_user_role()
        user_email = login.get_user_email()
        user_departamento = login.get_user_departamento()  # Obtener el departamento
        
        # Crear la ventana principal pasando email en lugar de username
        window = MainWindow(email=user_email, user_role=user_role, departamento=user_departamento)  # Corregido aquí
        
        # Configurar visibilidad de botones según el rol del usuario
        for button in window.findChildren(QToolButton):
            if user_role == "admin":
                # El administrador ve todos los botones
                button.setVisible(True)
            elif user_role == "recepcionista":
                # El recepcionista solo ve agregar y consultar datos
                if button.objectName() in ["Eliminar Datos", "Modificar Datos", "Administrar"]:
                    button.setVisible(False)
                else:
                    button.setVisible(True)
            else:  # usuario normal
                # Usuario normal solo ve consultar
                if button.objectName() not in ["Consultar Datos"]:
                    button.setVisible(False)  
        window.show()
        sys.exit(app.exec())
    else:
        sys.exit()

if __name__ == "__main__":
    main()